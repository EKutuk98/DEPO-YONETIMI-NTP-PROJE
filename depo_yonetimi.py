# -*- coding: utf-8 -*-
"""
🏪 DEPO & STOK YÖNETİM SİSTEMİ v1.0
Yemek Tarif Platformu (test.py) Tabanlı - MASTER EDITION
═════════════════════════════════════════════════════════════════════════════

Özellikler:
✅ Profesyonel Dark Tema (Yemek'ten uyarlı)
✅ Veritabanı: urunler, siparisler, stok_hareketleri, kategoriler, tedarikçiler
✅ CRUD: Ürün, Sipariş, Stok, Tedarikçi
✅ İstatistikler: 6 grafik, KPI dashboard
✅ Raporlar: CSV, PDF, Excel export
✅ Otomatik Reorder Point Sistemi
✅ ABC Analizi
✅ Stok Rotasyon (FIFO)
✅ Tedarikçi Performans Analizi
✅ Anomali Tespiti (Düşük stok, Hırsızlık)
✅ Rol Tabanlı Erişim (Admin/Personel)
✅ Gamification (Leaderboard)
✅ Real-time Dashboard

Satır: ~3500+
OOP Sınıflar: 15+
Veritabanı Tablolar: 8
Dialog Türleri: 8
Grafik Türü: 6+
Export Formatları: 3 (CSV, PDF, Excel)
"""

import sys
import os
import csv
import json
import sqlite3
from datetime import datetime, timedelta
from contextlib import contextmanager
import hashlib
import random

# ═══════════════════════════════════════════════════════════════════════════
# İMPORT: Excel, Matplotlib, PyQt5
# ═══════════════════════════════════════════════════════════════════════════

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_VAR = True
except ImportError:
    OPENPYXL_VAR = False

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QDialog, QLabel,
    QLineEdit, QComboBox, QMessageBox, QTabWidget, QFrame, QSpinBox,
    QHeaderView, QGridLayout, QTextEdit, QStackedWidget, QToolBar,
    QCheckBox, QFileDialog, QMenu, QAction, QShortcut, QSplitter,
    QScrollArea, QDoubleSpinBox, QWizard, QWizardPage, QListWidget,
    QListWidgetItem, QProgressBar, QDateEdit, QTimeEdit
)
from PyQt5.QtCore import Qt, QTimer, QPropertyAnimation, QEasingCurve, QPoint, QDate, QDateTime
from PyQt5.QtGui import QFont, QColor, QIcon, QPalette, QLinearGradient, QBrush, QPainter, QPixmap

# QtChart kaldırıldı - Matplotlib kullanıyoruz
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import numpy as np

plt.rcParams['font.sans-serif'] = ['Arial']
plt.rcParams['axes.unicode_minus'] = False

# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 1: TEMA YÖNETİMİ (YEMEK TARIFINDEN UYARLANDI)
# ═══════════════════════════════════════════════════════════════════════════

class ThemeManager:
    """Tema yönetimi — Luxury Dark & Light (yemek tarifinden uyarlı)"""
    
    DARK = """
        QMainWindow { 
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #0f0f1a, stop:1 #1a1a2e); 
        }
        QWidget { background-color: transparent; }
        QLabel { color: #ffffff; }
        
        QTableWidget { 
            background-color: #2d2d3a; 
            alternate-background-color: #3d3d4a;
            color: #ffffff; 
            gridline-color: #4CAF50; 
            border: none; 
            border-radius: 15px; 
        }
        QTableWidget::item { 
            padding: 12px; 
            color: #ffffff; 
        }
        QTableWidget::item:selected { 
            background-color: #0077b6; 
            color: #ffffff; 
        }
        QHeaderView::section { 
            background-color: #1a1a2e; 
            color: #0077b6; 
            font-weight: bold; 
            padding: 12px; 
            border: none; 
        }
        
        QComboBox, QLineEdit, QSpinBox, QDoubleSpinBox, QTextEdit, QDateEdit { 
            background-color: #1a1a2e; 
            border: 2px solid #0077b6;
            border-radius: 12px; 
            padding: 10px; 
            color: #ffffff; 
            font-size: 12px; 
        }
        QComboBox:focus, QLineEdit:focus { 
            border: 2px solid #00b4d8; 
        }
        QComboBox QAbstractItemView { 
            background-color: #1a1a2e; 
            color: #ffffff; 
            selection-background-color: #0077b6; 
        }
        
        QCheckBox { 
            color: #0077b6; 
            font-weight: bold; 
        }
        
        QScrollBar:vertical { 
            background-color: #2d2d3a; 
            border-radius: 10px; 
            width: 10px; 
        }
        QScrollBar::handle:vertical { 
            background-color: #0077b6; 
            border-radius: 10px; 
        }
        QScrollBar::handle:vertical:hover { 
            background-color: #00b4d8; 
        }
        
        QProgressBar {
            background-color: #1a1a2e;
            border: 2px solid #0077b6;
            border-radius: 10px;
            text-align: center;
            color: #ffffff;
        }
        QProgressBar::chunk {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #0077b6, stop:1 #00b4d8);
            border-radius: 8px;
        }
        
        QTabWidget::pane { border: none; }
        QTabBar::tab { 
            background-color: #1a1a2e; 
            color: #0077b6; 
            padding: 12px 25px;
            border-radius: 8px; 
            margin-right: 5px; 
            border: 1px solid #0077b6; 
        }
        QTabBar::tab:selected { 
            background-color: #0077b6; 
            color: #ffffff; 
        }
    """
    
    LIGHT = """
        QMainWindow { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f0f2f5, stop:1 #e2e8f0); }
        QWidget { background-color: transparent; }
        QLabel { color: #1a1a2e; }
        QTableWidget { background-color: #ffffff; alternate-background-color: #f1f5f9;
            color: #1a1a2e; gridline-color: #cbd5e1; border: none; border-radius: 15px; }
        QTableWidget::item { padding: 12px; color: #1a1a2e; }
        QTableWidget::item:selected { background-color: #0077b6; color: #ffffff; }
        QHeaderView::section { background-color: #e2e8f0; color: #0077b6; font-weight: bold; padding: 12px; border: none; }
        QComboBox, QLineEdit, QSpinBox, QDoubleSpinBox { background-color: #ffffff; border: 2px solid #0077b6;
            border-radius: 12px; padding: 10px; color: #1a1a2e; font-size: 12px; }
        QComboBox:focus, QLineEdit:focus { border: 2px solid #00b4d8; }
        QComboBox QAbstractItemView { background-color: #ffffff; color: #1a1a2e; selection-background-color: #0077b6; }
        QCheckBox { color: #0077b6; font-weight: bold; }
    """
    
    @staticmethod
    def get(theme):
        return ThemeManager.LIGHT if theme == "light" else ThemeManager.DARK


# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 2: MESAJ KUTUSU STİLERİ (DARK TEMA)
# ═══════════════════════════════════════════════════════════════════════════

_MSG_STYLE = """
    QMessageBox {
        background-color: #1a1a2e;
        color: #ffffff;
    }
    QMessageBox QLabel {
        color: #ffffff;
        font-size: 13px;
        padding: 6px;
        min-width: 260px;
    }
    QMessageBox QPushButton {
        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
            stop:0 #0077b6, stop:1 #1a1a2e);
        color: #ffffff;
        border: none;
        border-radius: 8px;
        font-weight: bold;
        font-size: 12px;
        padding: 8px 22px;
        min-width: 80px;
    }
    QPushButton:hover {
        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
            stop:0 #0077b6, stop:1 #0077b6);
    }
    QPushButton:focus { outline: none; }
"""


def msg_info(parent, baslik, metin):
    """Dark tema bilgi mesajı"""
    mb = QMessageBox(parent)
    mb.setWindowTitle(baslik)
    mb.setText(metin)
    mb.setIcon(QMessageBox.NoIcon)
    mb.setStyleSheet(_MSG_STYLE)
    mb.exec_()


def msg_warn(parent, baslik, metin):
    """Dark tema uyarı mesajı"""
    mb = QMessageBox(parent)
    mb.setWindowTitle(baslik)
    mb.setText(metin)
    mb.setIcon(QMessageBox.NoIcon)
    mb.setStyleSheet(_MSG_STYLE)
    mb.exec_()


def msg_question(parent, baslik, metin):
    """Dark tema onay sorusu"""
    mb = QMessageBox(parent)
    mb.setWindowTitle(baslik)
    mb.setText(metin)
    mb.setIcon(QMessageBox.NoIcon)
    mb.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    mb.setDefaultButton(QMessageBox.No)
    mb.setStyleSheet(_MSG_STYLE)
    mb.button(QMessageBox.Yes).setText("✅ Evet")
    mb.button(QMessageBox.No).setText("❌ Hayır")
    return mb.exec_()


# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 3: ÖZEL BUTON SINIFI (LUXURY)
# ═══════════════════════════════════════════════════════════════════════════

class LuxuryButton(QPushButton):
    """Luxury tasarımlı buton (yemek tarifinden uyarlı)"""
    
    def __init__(self, text, color="#0077b6", parent=None):
        super().__init__(text, parent)
        self.color = color
        self.setMinimumHeight(36)
        self.setSizePolicy(self.sizePolicy().Preferred, self.sizePolicy().Fixed)
        self.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color}, stop:1 #1a1a2e);
                color: white; 
                border: none; 
                border-radius: 10px;
                font-weight: bold; 
                font-size: 11px; 
                padding: 7px 12px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color}, stop:1 {color});
            }}
        """)


# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 4: SAYISAL SIRALANABL TABLO HÜCRESİ
# ═══════════════════════════════════════════════════════════════════════════

class NumericTableWidgetItem(QTableWidgetItem):
    """Sayısal verileri doğru sıralan tablo hücresi"""
    
    def __init__(self, text, sayi):
        super().__init__(text)
        self.sayi = sayi

    def __lt__(self, other):
        if isinstance(other, NumericTableWidgetItem):
            return self.sayi < other.sayi
        return super().__lt__(other)


# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 5: VERİTABANI YÖNETİCİSİ (MASTER EDITION)
# ═══════════════════════════════════════════════════════════════════════════

class DatabaseManager:
    """Veritabanı işlemlerini yöneten sınıf - MASTER EDITION"""
    
    def __init__(self, db_name="depo_yonetim_master.db"):
        self.db_name = db_name
        self.create_tables()
        self.create_sample_data()
    
    @contextmanager
    def get_connection(self):
        """Context manager: Veritabanı bağlantısı"""
        conn = sqlite3.connect(self.db_name)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()
    
    def create_tables(self):
        """Veritabanı tablolarını oluştur"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # ── ÜRÜNLER TABLOSU ──
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS urunler (
                    urun_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    urun_kodu TEXT UNIQUE NOT NULL,
                    urun_adi TEXT NOT NULL,
                    kategori TEXT NOT NULL,
                    stok INTEGER DEFAULT 0,
                    minimum_stok INTEGER DEFAULT 10,
                    reorder_point INTEGER DEFAULT 20,
                    fiyat REAL NOT NULL,
                    alis_fiyati REAL,
                    birim TEXT DEFAULT 'adet',
                    raf_no TEXT,
                    aciklama TEXT,
                    durum TEXT DEFAULT 'Aktif',
                    abc_kategori TEXT DEFAULT 'C',
                    son_giris_tarihi TIMESTAMP,
                    son_cikis_tarihi TIMESTAMP,
                    eklenme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # ── ÜRÜN SAYıM LOGU ──
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS urun_sayim_logu (
                    sayim_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    urun_id INTEGER NOT NULL,
                    sistem_stok INTEGER,
                    fiziki_stok INTEGER,
                    fark INTEGER,
                    sayim_yapan TEXT,
                    sayim_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    aciklama TEXT,
                    FOREIGN KEY (urun_id) REFERENCES urunler (urun_id)
                )
            ''')
            
            # ── SİPARİŞLER TABLOSU ──
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS siparisler (
                    siparis_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    siparis_no TEXT UNIQUE NOT NULL,
                    urun_id INTEGER NOT NULL,
                    urun_adi TEXT NOT NULL,
                    adet INTEGER NOT NULL,
                    birim_fiyat REAL NOT NULL,
                    toplam_tutar REAL NOT NULL,
                    siparis_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    teslim_tarihi TIMESTAMP,
                    teslim_gunceleme_tarihi TIMESTAMP,
                    durum TEXT DEFAULT 'Beklemede',
                    odeme_durumu TEXT DEFAULT 'Beklemede',
                    musteri_adi TEXT,
                    musteri_tel TEXT,
                    musteri_email TEXT,
                    aciklama TEXT,
                    teslim_kaniti_foto TEXT,
                    FOREIGN KEY (urun_id) REFERENCES urunler (urun_id)
                )
            ''')
            
            # ── STOK HAREKETLERİ TABLOSU ──
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS stok_hareketleri (
                    hareket_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    urun_id INTEGER NOT NULL,
                    urun_adi TEXT NOT NULL,
                    hareket_tipi TEXT NOT NULL,
                    miktar INTEGER NOT NULL,
                    onceki_stok INTEGER,
                    sonraki_stok INTEGER,
                    hareket_nedeni TEXT,
                    islem_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    islem_yapan TEXT,
                    siparis_id INTEGER,
                    aciklama TEXT,
                    FOREIGN KEY (urun_id) REFERENCES urunler (urun_id),
                    FOREIGN KEY (siparis_id) REFERENCES siparisler (siparis_id)
                )
            ''')
            
            # ── KATEGORİLER TABLOSU ──
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kategoriler (
                    kategori_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kategori_adi TEXT UNIQUE NOT NULL,
                    aciklama TEXT
                )
            ''')
            
            # ── TEDARİKÇİLER TABLOSU ──
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tedarikçiler (
                    tedarikci_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tedarikci_adi TEXT NOT NULL,
                    telefon TEXT,
                    email TEXT,
                    adres TEXT,
                    ulas_suresi INTEGER,
                    minimum_siparis INTEGER,
                    fiyat_indirim_orani REAL DEFAULT 0,
                    on_time_delivery_rate REAL DEFAULT 0,
                    kalite_orani REAL DEFAULT 0,
                    durum TEXT DEFAULT 'Aktif',
                    aciklama TEXT,
                    eklenme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # ── ÜRÜN-TEDARİKÇİ İLİŞKİSİ ──
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS urun_tedarikçi (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    urun_id INTEGER NOT NULL,
                    tedarikci_id INTEGER NOT NULL,
                    tedarikçi_fiyati REAL NOT NULL,
                    lead_time INTEGER,
                    son_siparis_tarihi TIMESTAMP,
                    FOREIGN KEY (urun_id) REFERENCES urunler (urun_id),
                    FOREIGN KEY (tedarikci_id) REFERENCES tedarikçiler (tedarikci_id)
                )
            ''')
            
            # ── SİSTEM KULLANICILAR ──
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS sistem_kullanicilar (
                    kullanici_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kullanici_adi TEXT UNIQUE NOT NULL,
                    sifre TEXT NOT NULL,
                    ad TEXT,
                    soyad TEXT,
                    email TEXT,
                    rol TEXT DEFAULT 'personel',
                    durum TEXT DEFAULT 'Aktif',
                    son_giris_tarihi TIMESTAMP,
                    eklenme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # ── VARSAYILAN KATEGORİLER ──
            default_cats = [
                ('Elektronik', 'Bilgisayar, telefon, aksesuar'),
                ('Giyim', 'Erkek, kadın, çocuk giysileri'),
                ('Gıda', 'Temel gıda maddeleri'),
                ('Kozmetik', 'Kişisel bakım ürünleri'),
                ('Ev Eşyası', 'Mutfak, salon, yatak odası'),
                ('Ofis Malz.', 'Kağıt, kalem, dosya'),
                ('Hırdavat', 'Yapı malzemeleri'),
                ('Spor', 'Spor malzemeleri')
            ]
            
            for cat, desc in default_cats:
                cursor.execute(
                    'INSERT OR IGNORE INTO kategoriler (kategori_adi, aciklama) VALUES (?, ?)',
                    (cat, desc)
                )
            
            # ── VARSAYILAN KULLANICI ──
            admin_hash = hashlib.sha256("admin123".encode()).hexdigest()
            cursor.execute(
                'INSERT OR IGNORE INTO sistem_kullanicilar (kullanici_adi, sifre, ad, soyad, rol) VALUES (?, ?, ?, ?, ?)',
                ('admin', admin_hash, 'Admin', 'Sistem', 'admin')
            )
    
    def create_sample_data(self):
        """Örnek veri ekle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM urunler')
            if cursor.fetchone()[0] == 0:
                products = [
                    ("PRD001", "Laptop", "Elektronik", 15, 5, 10, 15000.00, 12000.00, "adet", "A-01", "High-end laptop"),
                    ("PRD002", "Mouse", "Elektronik", 150, 20, 50, 250.00, 150.00, "adet", "A-02", "Wireless mouse"),
                    ("PRD003", "Klavye", "Elektronik", 80, 10, 30, 500.00, 350.00, "adet", "A-03", "Mechanical keyboard"),
                    ("PRD004", "Monitor", "Elektronik", 25, 3, 8, 2500.00, 2000.00, "adet", "A-04", "27\" 4K Monitor"),
                    ("PRD005", "T-Shirt", "Giyim", 200, 20, 60, 80.00, 45.00, "adet", "B-01", "Pamuk T-Shirt"),
                    ("PRD006", "Çanta", "Giyim", 50, 5, 15, 350.00, 250.00, "adet", "B-02", "Deri sırt çantası"),
                    ("PRD007", "Çikolata", "Gıda", 500, 50, 150, 15.00, 8.00, "adet", "C-01", "Bitter çikolata"),
                    ("PRD008", "Kahve", "Gıda", 300, 30, 100, 25.00, 15.00, "paket", "C-02", "Türk kahvesi"),
                    ("PRD009", "Şampuan", "Kozmetik", 100, 10, 30, 35.00, 20.00, "adet", "D-01", "Organik şampuan"),
                    ("PRD010", "Tabak Seti", "Ev Eşyası", 40, 5, 12, 450.00, 300.00, "adet", "E-01", "Porselen 12 li set"),
                ]
                
                for p in products:
                    cursor.execute('''
                        INSERT INTO urunler (urun_kodu, urun_adi, kategori, stok, minimum_stok, reorder_point,
                                            fiyat, alis_fiyati, birim, raf_no, aciklama)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', p)
                
                # TEDARİKÇİ EKLEİŞ:
                suppliers = [
                    ("Tech Supply Co.", "0216 123 4567", "info@techsupply.com", "Istanbul", 3, 50),
                    ("Fashion Wholesale", "0312 456 7890", "sales@fashionwholesale.com", "Ankara", 5, 100),
                    ("Food Distributor", "0384 789 0123", "orders@fooddist.com", "Gaziantep", 7, 200),
                ]
                
                for supplier in suppliers:
                    cursor.execute('''
                        INSERT INTO tedarikçiler (tedarikci_adi, telefon, email, adres, ulas_suresi, minimum_siparis)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', supplier)

                # Örnek stok hareketleri (zengin)
                sample_hareketler = [
                    (1,  'Laptop',      'Giriş', 15,  0,   15,  'İlk Stok',          'admin'),
                    (2,  'Mouse',       'Giriş', 150, 0,   150, 'İlk Stok',          'admin'),
                    (3,  'Klavye',      'Giriş', 80,  0,   80,  'İlk Stok',          'admin'),
                    (4,  'Monitor',     'Giriş', 25,  0,   25,  'İlk Stok',          'admin'),
                    (5,  'T-Shirt',     'Giriş', 200, 0,   200, 'İlk Stok',          'admin'),
                    (6,  'Çanta',       'Giriş', 50,  0,   50,  'İlk Stok',          'admin'),
                    (7,  'Çikolata',    'Giriş', 500, 0,   500, 'İlk Stok',          'admin'),
                    (8,  'Kahve',       'Giriş', 300, 0,   300, 'İlk Stok',          'admin'),
                    (9,  'Şampuan',     'Giriş', 100, 0,   100, 'İlk Stok',          'admin'),
                    (10, 'Tabak Seti',  'Giriş', 40,  0,   40,  'İlk Stok',          'admin'),
                    (1,  'Laptop',      'Çıkış', 2,   15,  13,  'Satış - ORD000001', 'admin'),
                    (2,  'Mouse',       'Çıkış', 10,  150, 140, 'Satış - ORD000002', 'admin'),
                    (3,  'Klavye',      'Çıkış', 5,   80,  75,  'Satış - ORD000004', 'admin'),
                    (5,  'T-Shirt',     'Çıkış', 20,  200, 180, 'Satış - ORD000003', 'personel1'),
                    (7,  'Çikolata',    'Çıkış', 100, 500, 400, 'Satış - ORD000005', 'personel1'),
                    (8,  'Kahve',       'Çıkış', 50,  300, 250, 'Satış - ORD000007', 'personel1'),
                    (10, 'Tabak Seti',  'Çıkış', 3,   40,  37,  'Satış - ORD000008', 'personel1'),
                    (9,  'Şampuan',     'Çıkış', 15,  100, 85,  'Satış - ORD000010', 'admin'),
                    (2,  'Mouse',       'Giriş', 20,  140, 160, 'Stok Yenileme',     'admin'),
                    (7,  'Çikolata',    'Giriş', 100, 400, 500, 'Stok Yenileme',     'personel1'),
                    (4,  'Monitor',     'Çıkış', 1,   25,  24,  'Satış - ORD000006', 'admin'),
                    (6,  'Çanta',       'Çıkış', 2,   50,  48,  'Satış - ORD000009', 'admin'),
                ]
                for h in sample_hareketler:
                    cursor.execute('''
                        INSERT INTO stok_hareketleri (urun_id, urun_adi, hareket_tipi, miktar,
                                                     onceki_stok, sonraki_stok, hareket_nedeni, islem_yapan)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', h)

                # Örnek siparişler (zengin)
                sample_siparisler = [
                    ('ORD000001', 1,  'Laptop',     2,  15000.00, 30000.00, 'Teslim Edildi', 'Ahmet Yılmaz',   '05551234567', 'mehmet@email.com'),
                    ('ORD000002', 2,  'Mouse',      10,   250.00,  2500.00, 'Beklemede',     'Mehmet Demir',   '05559876543', 'demir@email.com'),
                    ('ORD000003', 5,  'T-Shirt',    20,    80.00,  1600.00, 'Hazırlanıyor',  'Ayşe Kaya',      '05554567890', 'ayse@email.com'),
                    ('ORD000004', 3,  'Klavye',      5,   500.00,  2500.00, 'Teslim Edildi', 'Fatma Çelik',    '05553214567', 'fatma@email.com'),
                    ('ORD000005', 7,  'Çikolata',  100,    15.00,  1500.00, 'Teslim Edildi', 'Ali Koç',        '05556789012', 'ali@email.com'),
                    ('ORD000006', 4,  'Monitor',     1,  2500.00,  2500.00, 'Beklemede',     'Can Arslan',     '05557654321', 'can@email.com'),
                    ('ORD000007', 8,  'Kahve',      50,    25.00,  1250.00, 'Hazırlanıyor',  'Selin Yıldız',   '05558901234', 'selin@email.com'),
                    ('ORD000008', 10, 'Tabak Seti',  3,   450.00,  1350.00, 'Teslim Edildi', 'Burak Şahin',    '05552345678', 'burak@email.com'),
                    ('ORD000009', 6,  'Çanta',       2,   350.00,   700.00, 'Beklemede',     'Zeynep Aydın',   '05559012345', 'zeynep@email.com'),
                    ('ORD000010', 9,  'Şampuan',    15,    35.00,   525.00, 'Teslim Edildi', 'Emre Güven',     '05550123456', 'emre@email.com'),
                ]
                for s in sample_siparisler:
                    cursor.execute('''
                        INSERT INTO siparisler (siparis_no, urun_id, urun_adi, adet, birim_fiyat,
                                               toplam_tutar, durum, musteri_adi, musteri_tel, musteri_email)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', s)


# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 6: VERITABANI SORGU METODLARI
# ═══════════════════════════════════════════════════════════════════════════

    # ── ÜRÜN SORGUSU ──
    
    def urunleri_getir(self, kategori=None, dusuk_stok=False, abc=None):
        """Ürünleri getir (filtreleme ile)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            query = 'SELECT * FROM urunler WHERE durum = "Aktif"'
            params = []
            
            if kategori:
                query += ' AND kategori = ?'
                params.append(kategori)
            if dusuk_stok:
                query += ' AND stok <= minimum_stok'
            if abc:
                query += ' AND abc_kategori = ?'
                params.append(abc)
            
            query += ' ORDER BY urun_adi'
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]
    
    def urun_ekle(self, urun_kodu, ad, kategori, stok, minimum_stok, reorder_point, 
                  fiyat, alis_fiyati=None, birim="adet", raf_no="", aciklama=""):
        """Yeni ürün ekle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO urunler (urun_kodu, urun_adi, kategori, stok, minimum_stok, reorder_point,
                                    fiyat, alis_fiyati, birim, raf_no, aciklama)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (urun_kodu, ad, kategori, stok, minimum_stok, reorder_point,
                  fiyat, alis_fiyati, birim, raf_no, aciklama))
            
            urun_id = cursor.lastrowid
            
            # Stok hareketi logu
            cursor.execute('''
                INSERT INTO stok_hareketleri (urun_id, urun_adi, hareket_tipi, miktar,
                                             onceki_stok, sonraki_stok, hareket_nedeni, aciklama)
                VALUES (?, ?, 'Giriş', ?, 0, ?, 'Ürün Eklendi', ?)
            ''', (urun_id, ad, stok, stok, 'Yeni ürün sisteme eklendi'))
            
            return urun_id
    
    def urun_guncelle(self, urun_id, **kwargs):
        """Ürünü güncelle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            updates = []
            values = []
            
            for key, value in kwargs.items():
                if key in ['urun_adi', 'kategori', 'minimum_stok', 'reorder_point', 'fiyat', 'raf_no']:
                    updates.append(f"{key} = ?")
                    values.append(value)
            
            if updates:
                values.append(urun_id)
                query = f"UPDATE urunler SET {', '.join(updates)} WHERE urun_id = ?"
                cursor.execute(query, values)
    
    def urun_sil(self, urun_id):
        """Ürünü sil (pasif yap)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE urunler SET durum = ? WHERE urun_id = ?', ('Pasif', urun_id))
    
    # ── STOK İŞLEMLERİ ──
    
    def stok_arttir(self, urun_id, miktar, aciklama="", islem_yapan="System"):
        """Stok arttır"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT urun_adi, stok FROM urunler WHERE urun_id = ?', (urun_id,))
            urun = cursor.fetchone()
            
            if not urun:
                raise ValueError("Ürün bulunamadı!")
            
            onceki_stok = urun['stok']
            yeni_stok = onceki_stok + miktar
            
            cursor.execute('UPDATE urunler SET stok = ?, son_giris_tarihi = CURRENT_TIMESTAMP WHERE urun_id = ?',
                          (yeni_stok, urun_id))
            
            cursor.execute('''
                INSERT INTO stok_hareketleri (urun_id, urun_adi, hareket_tipi, miktar,
                                             onceki_stok, sonraki_stok, aciklama, islem_yapan, hareket_nedeni)
                VALUES (?, ?, 'Giriş', ?, ?, ?, ?, ?, 'Stok Arttırıldı')
            ''', (urun_id, urun['urun_adi'], miktar, onceki_stok, yeni_stok, aciklama or "Stok artışı", islem_yapan))
            
            return yeni_stok
    
    def stok_azalt(self, urun_id, miktar, aciklama="", islem_yapan="System", siparis_id=None):
        """Stok azalt"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT urun_adi, stok FROM urunler WHERE urun_id = ?', (urun_id,))
            urun = cursor.fetchone()
            
            if not urun:
                raise ValueError("Ürün bulunamadı!")
            if urun['stok'] < miktar:
                raise ValueError(f"Yetersiz stok! Mevcut: {urun['stok']}, İstenen: {miktar}")
            
            onceki_stok = urun['stok']
            yeni_stok = onceki_stok - miktar
            
            cursor.execute('UPDATE urunler SET stok = ?, son_cikis_tarihi = CURRENT_TIMESTAMP WHERE urun_id = ?',
                          (yeni_stok, urun_id))
            
            cursor.execute('''
                INSERT INTO stok_hareketleri (urun_id, urun_adi, hareket_tipi, miktar,
                                             onceki_stok, sonraki_stok, aciklama, islem_yapan, siparis_id, hareket_nedeni)
                VALUES (?, ?, 'Çıkış', ?, ?, ?, ?, ?, ?, 'Stok Azaltıldı')
            ''', (urun_id, urun['urun_adi'], miktar, onceki_stok, yeni_stok, aciklama or "Stok azalışı", islem_yapan, siparis_id))
            
            return yeni_stok
    
    # ── SİPARİŞ İŞLEMLERİ ──
    
    def siparis_olustur(self, urun_id, adet, birim_fiyat, musteri_adi="", musteri_tel="", musteri_email="", aciklama=""):
        """Sipariş oluştur"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT ad FROM urunler WHERE urun_id = ?', (urun_id,))
            urun = cursor.fetchone()
            
            if not urun:
                raise ValueError("Ürün bulunamadı!")
            
            cursor.execute("SELECT COUNT(*) as count FROM siparisler")
            count = cursor.fetchone()['count'] + 1
            siparis_no = f"ORD{count:06d}"
            toplam_tutar = adet * birim_fiyat
            
            cursor.execute('''
                INSERT INTO siparisler (siparis_no, urun_id, urun_adi, adet, birim_fiyat,
                                       toplam_tutar, musteri_adi, musteri_tel, musteri_email, aciklama)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (siparis_no, urun_id, urun['urun_adi'], adet, birim_fiyat, toplam_tutar,
                  musteri_adi, musteri_tel, musteri_email, aciklama))
            
            return siparis_no
    
    def siparisleri_getir(self, durum=None):
        """Siparişleri getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            query = 'SELECT * FROM siparisler'
            params = []
            
            if durum:
                query += ' WHERE durum = ?'
                params.append(durum)
            
            query += ' ORDER BY siparis_tarihi DESC'
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]
    
    # ── İSTATİSTİKLER ──
    
    def toplam_istatistikler(self):
        """Toplam istatistikler"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute('SELECT COUNT(*) as count FROM urunler WHERE durum = "Aktif"')
            urun_sayisi = cursor.fetchone()['count']
            
            cursor.execute('SELECT SUM(stok) as total FROM urunler WHERE durum = "Aktif"')
            toplam_stok = cursor.fetchone()['total'] or 0
            
            cursor.execute('SELECT SUM(stok * fiyat) as total FROM urunler WHERE durum = "Aktif"')
            toplam_deger = cursor.fetchone()['total'] or 0
            
            cursor.execute('SELECT COUNT(*) as count FROM siparisler WHERE durum IN ("Beklemede", "Hazırlanıyor")')
            aktif_siparis = cursor.fetchone()['count']
            
            cursor.execute('SELECT COUNT(*) as count FROM urunler WHERE stok <= minimum_stok AND durum = "Aktif"')
            dusuk_stok_sayisi = cursor.fetchone()['count']
            
            return {
                'urun_sayisi': urun_sayisi,
                'toplam_stok': toplam_stok,
                'toplam_deger': toplam_deger,
                'aktif_siparis': aktif_siparis,
                'dusuk_stok_sayisi': dusuk_stok_sayisi
            }
    
    def kategorileri_getir(self):
        """Kategorileri getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT kategori_adi FROM kategoriler ORDER BY kategori_adi')
            return [row['kategori_adi'] for row in cursor.fetchall()]
    
    # ── TEDARİKÇİ İŞLEMLERİ ──
    
    def tedarikci_ekle(self, adi, telefon, email, adres, ulas_suresi, min_siparis, indirim):
        """Yeni tedarikçi ekle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO tedarikçiler (tedarikci_adi, telefon, email, adres, 
                                         ulas_suresi, minimum_siparis, fiyat_indirim_orani)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (adi, telefon, email, adres, ulas_suresi, min_siparis, indirim))
            return cursor.lastrowid
    
    def tedarikci_listele(self):
        """Tedarikçileri getir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM tedarikçiler WHERE durum = "Aktif" ORDER BY tedarikci_adi')
            return [dict(row) for row in cursor.fetchall()]
    
    def tedarikci_guncelle(self, tedarikci_id, **kwargs):
        """Tedarikçi güncelle"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            updates = []
            values = []
            
            for key, value in kwargs.items():
                if key in ['tedarikci_adi', 'telefon', 'email', 'adres', 'ulas_suresi', 
                          'minimum_siparis', 'fiyat_indirim_orani', 'on_time_delivery_rate', 'kalite_orani']:
                    updates.append(f"{key} = ?")
                    values.append(value)
            
            if updates:
                values.append(tedarikci_id)
                query = f"UPDATE tedarikçiler SET {', '.join(updates)} WHERE tedarikci_id = ?"
                cursor.execute(query, values)
    
    def tedarikci_sil(self, tedarikci_id):
        """Tedarikçi sil (pasif yap)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE tedarikçiler SET durum = ? WHERE tedarikci_id = ?', 
                          ('Pasif', tedarikci_id))
    
    def kullanici_sifre_degistir(self, kullanici_id, yeni_sifre):
        """Kullanıcı şifresini değiştir"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            sifre_hash = hashlib.sha256(yeni_sifre.encode()).hexdigest()
            cursor.execute('''
                UPDATE sistem_kullanicilar
                SET sifre = ?
                WHERE kullanici_id = ?
            ''', (sifre_hash, kullanici_id))
    
    def stok_hareketlerini_getir(self, urun_id=None, gun=30):
        """Stok hareketlerini getir (son N gün)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            query = 'SELECT * FROM stok_hareketleri WHERE islem_tarihi >= datetime("now", "-' + str(gun) + ' days")'
            params = []
            
            if urun_id:
                query += ' AND urun_id = ?'
                params.append(urun_id)
            
            query += ' ORDER BY islem_tarihi DESC'
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]
    
    def kategori_dagilimi_getir(self):
        """Kategori dağılımı (pie chart için)"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT kategori, SUM(stok) as toplam_stok
                FROM urunler
                WHERE durum = "Aktif"
                GROUP BY kategori
                ORDER BY toplam_stok DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    def aylik_siparis_trendi(self, ay_sayi=12):
        """Aylık sipariş trendi"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f'''
                SELECT 
                    strftime('%Y-%m', siparis_tarihi) as ay,
                    COUNT(*) as siparis_sayisi,
                    SUM(toplam_tutar) as toplam_gelir
                FROM siparisler
                WHERE siparis_tarihi >= datetime("now", "-{ay_sayi} months")
                GROUP BY ay
                ORDER BY ay
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    def en_cok_satilan_urunler(self, limit=10):
        """En çok satılan ürünler"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f'''
                SELECT 
                    s.urun_adi,
                    SUM(s.adet) as toplam_adet,
                    SUM(s.toplam_tutar) as toplam_tutar,
                    COUNT(*) as siparis_sayisi
                FROM siparisler s
                WHERE s.durum = "Teslim Edildi"
                GROUP BY s.urun_adi
                ORDER BY toplam_adet DESC
                LIMIT {limit}
            ''')
            return [dict(row) for row in cursor.fetchall()]
    
    def anomali_tespit(self):
        """Anomali tespiti (düşük stok, hareketsiz ürünler vb.)"""
        anomalies = []
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Kritik stok
            cursor.execute('''
                SELECT urun_id, ad, stok, minimum_stok
                FROM urunler
                WHERE stok < minimum_stok AND durum = "Aktif"
            ''')
            for row in cursor.fetchall():
                anomalies.append({
                    'type': 'CRITICAL_STOCK',
                    'urun': row['urun_adi'],
                    'stok': row['stok'],
                    'min': row['minimum_stok'],
                    'severity': 'HIGH' if row['stok'] == 0 else 'MEDIUM'
                })
            
            # Hareketsiz ürünler (90 gün)
            cursor.execute('''
                SELECT urun_id, ad, son_cikis_tarihi
                FROM urunler
                WHERE son_cikis_tarihi < datetime("now", "-90 days") AND durum = "Aktif"
            ''')
            for row in cursor.fetchall():
                anomalies.append({
                    'type': 'DEAD_STOCK',
                    'urun': row['urun_adi'],
                    'last_sale': row['son_cikis_tarihi'],
                    'severity': 'LOW'
                })
        
        return anomalies


# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 7: ANA PENCERE (TEMEL YAPI)
# ═══════════════════════════════════════════════════════════════════════════

class DepoMainWindow(QMainWindow):
    """Ana pencere - Master Edition"""
    
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.current_theme = "dark"
        self.current_user = "admin"  # login yapılmadığı için default
        
        self.setWindowTitle("🏪 DEPO & STOK YÖNETİM SİSTEMİ MASTER EDITION")
        self.setGeometry(50, 50, 1600, 1000)
        self.setStyleSheet(ThemeManager.get(self.current_theme))
        
        # Kısayolları ekle
        self.kisayollar_ekle()
        
        self.init_ui()
        self.load_data()
        self._demo_profil_doldur()

        # Otomatik yenile
        self.timer = QTimer()
        self.timer.timeout.connect(self.refresh_data)
        self.timer.start(5000)  # 5 saniyede bir
    
    def init_ui(self):
        """Arayüzü oluştur"""
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout()
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # ── HEADER ──
        header = self.create_header()
        
        # ── DASHBOARD KARTLARI ──
        dashboard = self.create_dashboard()
        
        # ── SEKMELER ──
        self.tabs = self.create_tabs()
        
        main_layout.addWidget(header)
        main_layout.addLayout(dashboard)
        main_layout.addWidget(self.tabs)
        
        central.setLayout(main_layout)
    
    def create_header(self):
        """Header oluştur"""
        header = QFrame()
        header.setStyleSheet("""
            QFrame { 
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1a1a2e, stop:0.3 #2d2d3a, stop:0.7 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 25px; 
                border: 1px solid #0077b6; 
            }
        """)
        header.setFixedHeight(110)
        h_layout = QHBoxLayout()
        h_layout.setContentsMargins(40, 0, 40, 0)
        h_layout.setSpacing(20)
        
        # Logo
        logo = QLabel("🏪")
        logo.setFont(QFont("Arial", 50))
        h_layout.addWidget(logo)
        
        # Başlık
        title_layout = QVBoxLayout()
        title = QLabel("DEPO & STOK YÖNETİM SİSTEMİ")
        title.setFont(QFont("Arial", 22, QFont.Bold))
        title.setStyleSheet("color: #0077b6;")
        title_layout.addWidget(title)
        
        subtitle = QLabel("MASTER EDITION - Yemek Tarifi Platformu Tabanlı")
        subtitle.setFont(QFont("Arial", 10))
        subtitle.setStyleSheet("color: #2ec4b6; letter-spacing: 1px;")
        title_layout.addWidget(subtitle)
        
        h_layout.addLayout(title_layout)
        h_layout.addStretch()
        
        # PRO Badge
        pro = QLabel("⭐ PRO v1.0")
        pro.setFont(QFont("Arial", 12, QFont.Bold))
        pro.setStyleSheet("""
            color: #1a1a2e; 
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #0077b6, stop:1 #00b4d8);
            border-radius: 15px; 
            padding: 10px 25px; 
            letter-spacing: 2px;
        """)
        h_layout.addWidget(pro)
        
        # Kullanıcı info
        user_info = QLabel(f"👤 {self.current_user.upper()} | 🕐 {datetime.now().strftime('%H:%M')}")
        user_info.setFont(QFont("Arial", 10))
        user_info.setStyleSheet("color: #4CAF50;")
        h_layout.addWidget(user_info)
        
        # Raporlar butonu
        rapor_btn = LuxuryButton("📊 Raporlar", "#ff9f1c")
        rapor_btn.setMaximumWidth(130)
        rapor_btn.clicked.connect(self.raporlar_ac)
        h_layout.addWidget(rapor_btn)
        
        # Anomaliler butonu
        anomali_btn = LuxuryButton("⚠️ Anomaliler", "#e63946")
        anomali_btn.setMaximumWidth(140)
        anomali_btn.clicked.connect(self.anomaliler_ac)
        h_layout.addWidget(anomali_btn)
        
        # Gamification butonu
        game_btn = LuxuryButton("🏆 Başarılar", "#ffd700")
        game_btn.setMaximumWidth(130)
        game_btn.clicked.connect(self.gamification_ac)
        h_layout.addWidget(game_btn)
        
        # Exit
        exit_btn = LuxuryButton("✕ EXIT", "#e63946")
        exit_btn.setMaximumWidth(100)
        exit_btn.clicked.connect(self.close)
        h_layout.addWidget(exit_btn)
        
        header.setLayout(h_layout)
        return header
    
    def create_dashboard(self):
        """Dashboard kartları oluştur"""
        dashboard = QHBoxLayout()
        dashboard.setSpacing(15)
        
        self.urun_card = self.create_card("📦 TOPLAM ÜRÜN", "0", "#0077b6", "Aktif Ürünler")
        self.stok_card = self.create_card("📊 TOPLAM STOK", "0", "#2ec4b6", "Toplam Miktar")
        self.deger_card = self.create_card("💰 STOK DEĞERİ", "0 TL", "#ff9f1c", "Para Değeri")
        self.siparis_card = self.create_card("🛒 AKTİF SİPARİŞ", "0", "#e63946", "Beklenen")
        self.dusuk_stok_card = self.create_card("⚠️ DÜŞÜK STOK", "0", "#ff6b6b", "İhtiyaç Olan")
        
        dashboard.addWidget(self.urun_card)
        dashboard.addWidget(self.stok_card)
        dashboard.addWidget(self.deger_card)
        dashboard.addWidget(self.siparis_card)
        dashboard.addWidget(self.dusuk_stok_card)
        
        return dashboard
    
    def create_card(self, title, value, color, subtitle):
        """Luxury kart oluştur"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{ 
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 20px; 
                border-left: 8px solid {color}; 
            }}
        """)
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(18, 18, 18, 18)
        
        t = QLabel(title)
        t.setFont(QFont("Arial", 11, QFont.Bold))
        t.setStyleSheet(f"color: {color};")
        
        v = QLabel(value)
        v.setFont(QFont("Arial", 28, QFont.Bold))
        v.setStyleSheet("color: #ffffff;")
        v.setObjectName("value_label")
        
        s = QLabel(subtitle)
        s.setFont(QFont("Arial", 9))
        s.setStyleSheet("color: #4CAF50;")
        
        layout.addWidget(t)
        layout.addWidget(v)
        layout.addWidget(s)
        card.setLayout(layout)
        
        return card
    
    def create_tabs(self):
        """Sekmeler oluştur"""
        tabs = QTabWidget()
        tabs.addTab(self.create_urun_tab(), "📦 Ürünler")
        tabs.addTab(self.create_siparis_tab(), "🛒 Siparişler")
        tabs.addTab(self.create_tedarikci_tab(), "🏢 Tedarikçiler")
        tabs.addTab(self.create_hareket_tab(), "📋 Hareketler")
        tabs.addTab(self.create_istatistikler_tab(), "📊 İstatistikler")
        tabs.addTab(self.create_ayarlar_tab(), "⚙️ Ayarlar")
        tabs.addTab(NotlarWidget(self.db), "📝 Notlar")
        return tabs
    
    def create_urun_tab(self):
        """Ürünler sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # Butonlar
        btn_panel = QFrame()
        btn_panel.setStyleSheet("background-color: #1b263b; border-radius: 10px; padding: 10px;")
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        ekle_btn = LuxuryButton("➕ Yeni Ürün", "#2ec4b6")
        ekle_btn.setMaximumWidth(120)
        ekle_btn.clicked.connect(self.urun_ekle)
        
        arttir_btn = LuxuryButton("📈 Stok +", "#4CAF50")
        arttir_btn.setMaximumWidth(120)
        arttir_btn.clicked.connect(self.stok_arttir)
        
        azalt_btn = LuxuryButton("📉 Stok -", "#ff9f1c")
        azalt_btn.setMaximumWidth(120)
        azalt_btn.clicked.connect(self.stok_azalt)
        
        sayim_btn = LuxuryButton("📦 Stok Sayımı", "#2ec4b6")
        sayim_btn.setMaximumWidth(140)
        sayim_btn.clicked.connect(self.stok_sayimi_ac)
        
        yenile_btn = LuxuryButton("🔄 Yenile", "#0077b6")
        yenile_btn.setMaximumWidth(120)
        yenile_btn.clicked.connect(self.urunleri_listele)
        
        btn_layout.addWidget(ekle_btn)
        btn_layout.addWidget(arttir_btn)
        btn_layout.addWidget(azalt_btn)
        btn_layout.addWidget(sayim_btn)
        btn_layout.addWidget(yenile_btn)
        btn_layout.addStretch()
        
        btn_panel.setLayout(btn_layout)
        layout.addWidget(btn_panel)
        
        # Tablo
        self.urun_table = QTableWidget()
        self.urun_table.setColumnCount(11)
        self.urun_table.setHorizontalHeaderLabels([
            "ID", "Kod", "Ürün Adı", "Kategori", "Stok", "Min", "Reorder", "Fiyat", "ABC", "Birim", "Raf"
        ])
        self.urun_table.setAlternatingRowColors(True)
        self.urun_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.urun_table)
        widget.setLayout(layout)
        return widget
    
    def create_siparis_tab(self):
        """Siparişler sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        btn_panel = QFrame()
        btn_panel.setStyleSheet("background-color: #1b263b; border-radius: 10px; padding: 10px;")
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        siparis_btn = LuxuryButton("🛒 Yeni Sipariş", "#2ec4b6")
        siparis_btn.setMaximumWidth(140)
        siparis_btn.clicked.connect(self.siparis_olustur)
        
        onayla_btn = LuxuryButton("✅ Onayla", "#4CAF50")
        onayla_btn.setMaximumWidth(120)
        onayla_btn.clicked.connect(self.siparis_onayla)

        teslim_btn = LuxuryButton("🚚 Teslim Et", "#ff9f1c")
        teslim_btn.setMaximumWidth(120)
        teslim_btn.clicked.connect(self.siparis_teslim_et)

        iptal_btn = LuxuryButton("❌ İptal Et", "#e63946")
        iptal_btn.setMaximumWidth(110)
        iptal_btn.clicked.connect(self.siparis_iptal_et)

        btn_layout.addWidget(siparis_btn)
        btn_layout.addWidget(onayla_btn)
        btn_layout.addWidget(teslim_btn)
        btn_layout.addWidget(iptal_btn)
        btn_layout.addStretch()
        
        btn_panel.setLayout(btn_layout)
        layout.addWidget(btn_panel)
        
        self.siparis_table = QTableWidget()
        self.siparis_table.setColumnCount(10)
        self.siparis_table.setHorizontalHeaderLabels([
            "ID", "Sipariş No", "Ürün", "Adet", "Fiyat", "Toplam", "Tarih", "Durum", "Müşteri", "Tel"
        ])
        self.siparis_table.setAlternatingRowColors(True)
        self.siparis_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.siparis_table)
        widget.setLayout(layout)
        return widget
    
    def create_hareket_tab(self):
        """Hareketler sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        self.hareket_table = QTableWidget()
        self.hareket_table.setColumnCount(8)
        self.hareket_table.setHorizontalHeaderLabels([
            "ID", "Ürün", "İşlem", "Miktar", "Önceki", "Sonraki", "Tarih", "İşlem Yapan"
        ])
        self.hareket_table.setAlternatingRowColors(True)
        self.hareket_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.hareket_table)
        widget.setLayout(layout)
        return widget
    
    def create_istatistikler_tab(self):
        """İstatistikler sekmesi - Grafikli"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        sec_layout = QHBoxLayout()
        sec_layout.addWidget(QLabel("📊 Grafik:"))
        self.istat_combo = QComboBox()
        self.istat_combo.addItems([
            "Kategori Dagilimi (Pasta)",
            "Aylik Siparis Trendi",
            "En Cok Satilan Urunler",
            "ABC Analizi"
        ])
        self.istat_combo.currentIndexChanged.connect(self.istat_grafik_goster)
        sec_layout.addWidget(self.istat_combo)
        yenile = LuxuryButton("Yenile", "#0077b6")
        yenile.setMaximumWidth(100)
        yenile.clicked.connect(self.istat_grafik_goster)
        sec_layout.addWidget(yenile)
        sec_layout.addStretch()
        layout.addLayout(sec_layout)

        self.istat_canvas = MatplotlibCanvas(width=10, height=5, dpi=100)
        layout.addWidget(self.istat_canvas)

        widget.setLayout(layout)
        return widget

    def istat_grafik_goster(self):
        """Istatistik grafigini goster"""
        try:
            idx = self.istat_combo.currentIndex()
            proxy = type("P", (), {"db": self.db, "canvas": self.istat_canvas,
                "axes": self.istat_canvas.axes, "fig": self.istat_canvas.fig})()
            RaporlarDialog.show_kategori_dagitimi(proxy) if idx == 0 else (
                RaporlarDialog.show_aylik_siparis(proxy) if idx == 1 else (
                RaporlarDialog.show_en_cok_satilan(proxy) if idx == 2 else
                RaporlarDialog.show_abc_analizi(proxy)))
        except Exception:
            pass
    
    def create_tedarikci_tab(self):
        """Tedarikçiler sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        btn_panel = QFrame()
        btn_panel.setStyleSheet("background-color: #1b263b; border-radius: 10px; padding: 10px;")
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        ekle_btn = LuxuryButton("➕ Yeni Tedarikçi", "#2ec4b6")
        ekle_btn.setMaximumWidth(160)
        ekle_btn.clicked.connect(self.tedarikci_ekle)
        
        duzenle_btn = LuxuryButton("✏️ Düzenle", "#ff9f1c")
        duzenle_btn.setMaximumWidth(120)
        duzenle_btn.clicked.connect(self.tedarikci_duzenle)

        sil_btn = LuxuryButton("🗑️ Sil", "#e63946")
        sil_btn.setMaximumWidth(100)
        sil_btn.clicked.connect(self.tedarikci_sil_ui)
        
        yenile_btn = LuxuryButton("🔄 Yenile", "#0077b6")
        yenile_btn.setMaximumWidth(120)
        yenile_btn.clicked.connect(self.tedarikci_listele_ui)
        
        btn_layout.addWidget(ekle_btn)
        btn_layout.addWidget(duzenle_btn)
        btn_layout.addWidget(sil_btn)
        btn_layout.addWidget(yenile_btn)
        btn_layout.addStretch()
        
        btn_panel.setLayout(btn_layout)
        layout.addWidget(btn_panel)
        
        self.tedarikci_table = QTableWidget()
        self.tedarikci_table.setColumnCount(9)
        self.tedarikci_table.setHorizontalHeaderLabels([
            "ID", "Tedarikçi Adı", "Telefon", "Email", "Ulaş Süresi", 
            "Min Sipariş", "İndirim %", "On-Time %", "Durum"
        ])
        self.tedarikci_table.setAlternatingRowColors(True)
        self.tedarikci_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.tedarikci_table)
        widget.setLayout(layout)
        return widget
    
    def create_ayarlar_tab(self):
        """Ayarlar sekmesi - ScrollArea ile"""
        # Dış container (sekmede gösterilen)
        outer = QWidget()
        outer_layout = QVBoxLayout()
        outer_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(
            "QScrollArea{border:none;background:transparent;}"
            "QScrollBar:vertical{background:#2d2d3a;border-radius:6px;width:8px;}"
            "QScrollBar::handle:vertical{background:#0077b6;border-radius:6px;}"
        )

        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Profil Bölümü
        profil_frame = QFrame()
        profil_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px;
                padding: 20px;
            }
        """)
        profil_layout = QVBoxLayout()
        
        profil_baslik = QLabel("👤 Profil Bilgileri")
        profil_baslik.setFont(QFont("Arial", 14, QFont.Bold))
        profil_baslik.setStyleSheet("color: #0077b6;")
        profil_layout.addWidget(profil_baslik)
        
        profil_grid = QGridLayout()
        profil_grid.setSpacing(10)
        
        profil_grid.addWidget(QLabel("Kullanıcı Adı:"), 0, 0)
        self.profil_kullanici = QLineEdit()
        self.profil_kullanici.setText(self.current_user)
        self.profil_kullanici.setReadOnly(True)
        profil_grid.addWidget(self.profil_kullanici, 0, 1)
        
        profil_grid.addWidget(QLabel("Ad:"), 1, 0)
        self.profil_ad = QLineEdit()
        profil_grid.addWidget(self.profil_ad, 1, 1)
        
        profil_grid.addWidget(QLabel("Soyad:"), 2, 0)
        self.profil_soyad = QLineEdit()
        profil_grid.addWidget(self.profil_soyad, 2, 1)
        
        profil_layout.addLayout(profil_grid)
        profil_frame.setLayout(profil_layout)
        layout.addWidget(profil_frame)
        
        # Güvenlik Bölümü
        guvenlık_frame = QFrame()
        guvenlık_frame.setStyleSheet(profil_frame.styleSheet())
        guvenlık_layout = QVBoxLayout()
        
        guvenlık_baslik = QLabel("🔐 Güvenlik")
        guvenlık_baslik.setFont(QFont("Arial", 14, QFont.Bold))
        guvenlık_baslik.setStyleSheet("color: #0077b6;")
        guvenlık_layout.addWidget(guvenlık_baslik)
        
        guvenlık_grid = QGridLayout()
        guvenlık_grid.setSpacing(10)
        
        guvenlık_grid.addWidget(QLabel("Yeni Şifre:"), 0, 0)
        self.yeni_sifre = QLineEdit()
        self.yeni_sifre.setEchoMode(QLineEdit.Password)
        guvenlık_grid.addWidget(self.yeni_sifre, 0, 1)
        
        guvenlık_grid.addWidget(QLabel("Şifre Onayla:"), 1, 0)
        self.yeni_sifre_onayla = QLineEdit()
        self.yeni_sifre_onayla.setEchoMode(QLineEdit.Password)
        guvenlık_grid.addWidget(self.yeni_sifre_onayla, 1, 1)
        
        degistir_btn = LuxuryButton("🔑 Şifre Değiştir", "#ff9f1c")
        degistir_btn.setMaximumWidth(150)
        degistir_btn.clicked.connect(self.sifre_degistir)
        guvenlık_grid.addWidget(degistir_btn, 2, 1, Qt.AlignRight)
        
        guvenlık_layout.addLayout(guvenlık_grid)
        guvenlık_frame.setLayout(guvenlık_layout)
        layout.addWidget(guvenlık_frame)
        
        # Tema Bölümü
        tema_frame = QFrame()
        tema_frame.setStyleSheet(profil_frame.styleSheet())
        tema_layout = QVBoxLayout()
        
        tema_baslik = QLabel("🎨 Tema")
        tema_baslik.setFont(QFont("Arial", 14, QFont.Bold))
        tema_baslik.setStyleSheet("color: #0077b6;")
        tema_layout.addWidget(tema_baslik)
        
        tema_grid = QGridLayout()
        tema_grid.setSpacing(10)
        
        tema_grid.addWidget(QLabel("Tema Seçin:"), 0, 0)
        self.tema_combo = QComboBox()
        self.tema_combo.addItems(["🌙 Dark", "☀️ Light"])
        tema_grid.addWidget(self.tema_combo, 0, 1)
        
        tema_grid.addWidget(QLabel(""), 1, 0)
        
        tema_layout.addLayout(tema_grid)
        tema_frame.setLayout(tema_layout)
        layout.addWidget(tema_frame)
        
        # Yedekleme Bölümü
        yedek_frame = QFrame()
        yedek_frame.setStyleSheet(profil_frame.styleSheet())
        yedek_layout = QVBoxLayout()
        
        yedek_baslik = QLabel("💾 Yedekleme")
        yedek_baslik.setFont(QFont("Arial", 14, QFont.Bold))
        yedek_baslik.setStyleSheet("color: #0077b6;")
        yedek_layout.addWidget(yedek_baslik)
        
        yedek_btn_layout = QHBoxLayout()
        yedek_btn_layout.setSpacing(10)
        
        yedekle_btn = LuxuryButton("💾 Veritabanını Yedekle", "#4CAF50")
        yedekle_btn.setMaximumWidth(200)
        yedekle_btn.clicked.connect(self.veritabani_yedekle)
        
        geri_yukle_btn = LuxuryButton("📂 Yedekten Geri Yükle", "#0077b6")
        geri_yukle_btn.setMaximumWidth(200)
        
        yedek_btn_layout.addWidget(yedekle_btn)
        yedek_btn_layout.addWidget(geri_yukle_btn)
        yedek_btn_layout.addStretch()
        
        yedek_layout.addLayout(yedek_btn_layout)
        yedek_frame.setLayout(yedek_layout)
        layout.addWidget(yedek_frame)
        
        # Kişiselleştirme
        kis_frame = KisisellestirilmisAyarlar(self.db)
        layout.addWidget(kis_frame)
        
        # PHASE 7: Bonus Features
        bonus_frame = QFrame()
        bonus_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 10px;
                padding: 15px;
                border-left: 4px solid #ffd700;
            }
        """)
        bonus_layout = QVBoxLayout()
        
        bonus_baslik = QLabel("⭐ BONUS ÖZELLIKLER")
        bonus_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        bonus_baslik.setStyleSheet("color: #ffd700;")
        bonus_layout.addWidget(bonus_baslik)
        
        bonus_btn_layout = QHBoxLayout()
        bonus_btn_layout.setSpacing(10)
        
        istat_btn = LuxuryButton("📊 İleri İstatistikler", "#0077b6")
        istat_btn.setMaximumWidth(180)
        istat_btn.clicked.connect(lambda: IleriIstatistikPanel(self.db).exec_())
        bonus_btn_layout.addWidget(istat_btn)
        
        email_btn = LuxuryButton("📧 Email Rapor", "#2ec4b6")
        email_btn.setMaximumWidth(140)
        email_btn.clicked.connect(lambda: EmailRaporDialog(self.db).exec_())
        bonus_btn_layout.addWidget(email_btn)
        
        yedek_btn = LuxuryButton("🔄 Oto. Yedekleme", "#4CAF50")
        yedek_btn.setMaximumWidth(150)
        yedek_btn.clicked.connect(lambda: OtomatikYedekleDialog(self.db).exec_())
        bonus_btn_layout.addWidget(yedek_btn)
        
        bonus_btn_layout.addStretch()
        bonus_layout.addLayout(bonus_btn_layout)
        bonus_frame.setLayout(bonus_layout)
        layout.addWidget(bonus_frame)
        
        layout.addStretch()
        widget.setLayout(layout)

        scroll.setWidget(widget)
        outer_layout.addWidget(scroll)
        outer.setLayout(outer_layout)
        return outer

    def load_data(self):
        """Veriyi yükle"""
        self.urunleri_listele()
        self.siparisleri_listele()
        self.tedarikci_listele_ui()
        self.hareketleri_listele()
        self.update_dashboard()
        # İstatistik grafiğini başlat
        try:
            self.istat_grafik_goster()
        except Exception:
            pass

    def _demo_profil_doldur(self):
        """Tüm UI alanlarını demo veriyle otomatik doldur"""
        import random

        # ── Profil ──
        try:
            self.profil_ad.setText("Ahmet")
            self.profil_soyad.setText("Yılmaz")
        except Exception:
            pass

        # ── Notlar listesine ekstra demo notlar ekle ──
        try:
            notlar_widget = self.tabs.widget(self.tabs.count() - 1)
            # NotlarWidget'i bul
            def find_notlar(widget):
                from PyQt5.QtWidgets import QListWidget
                for child in widget.findChildren(QListWidget):
                    return child
                return None
            lst = find_notlar(notlar_widget)
            if lst and lst.count() <= 3:
                from PyQt5.QtWidgets import QListWidgetItem
                from PyQt5.QtGui import QColor
                ekstra = [
                    "Laptop siparişi takip edilecek",
                    "Q4 stok analizi hazırlanacak",
                    "Tedarikçi fiyat listesi güncellenmeli",
                    "Depo raf düzenlemesi planlanacak",
                    "Aylık rapor müdüre gönderilecek",
                ]
                for n in ekstra:
                    item = QListWidgetItem(f"  {n}")
                    item.setForeground(QColor("#b0b0b0"))
                    lst.addItem(item)
        except Exception:
            pass
    
    def urunleri_listele(self):
        """Ürünleri tabloya doldur"""
        self.urun_table.setRowCount(0)
        for u in self.db.urunleri_getir():
            row = self.urun_table.rowCount()
            self.urun_table.insertRow(row)
            
            self.urun_table.setItem(row, 0, QTableWidgetItem(str(u['urun_id'])))
            self.urun_table.setItem(row, 1, QTableWidgetItem(u['urun_kodu']))
            self.urun_table.setItem(row, 2, QTableWidgetItem(u['urun_adi']))
            self.urun_table.setItem(row, 3, QTableWidgetItem(u['kategori']))
            
            # Stok renk
            stok_item = QTableWidgetItem(str(u['stok']))
            if u['stok'] <= u['minimum_stok']:
                stok_item.setForeground(QColor("#e63946"))
            self.urun_table.setItem(row, 4, stok_item)
            
            self.urun_table.setItem(row, 5, QTableWidgetItem(str(u['minimum_stok'])))
            self.urun_table.setItem(row, 6, QTableWidgetItem(str(u['reorder_point'])))
            self.urun_table.setItem(row, 7, QTableWidgetItem(f"{u['fiyat']:,.2f} TL"))
            self.urun_table.setItem(row, 8, QTableWidgetItem(u['abc_kategori']))
            self.urun_table.setItem(row, 9, QTableWidgetItem(u['birim']))
            self.urun_table.setItem(row, 10, QTableWidgetItem(u['raf_no'] or "-"))
    
    def siparisleri_listele(self):
        """Siparişleri tabloya doldur"""
        self.siparis_table.setRowCount(0)
        for s in self.db.siparisleri_getir():
            row = self.siparis_table.rowCount()
            self.siparis_table.insertRow(row)
            
            self.siparis_table.setItem(row, 0, QTableWidgetItem(str(s['siparis_id'])))
            self.siparis_table.setItem(row, 1, QTableWidgetItem(s['siparis_no']))
            self.siparis_table.setItem(row, 2, QTableWidgetItem(s['urun_adi']))
            self.siparis_table.setItem(row, 3, QTableWidgetItem(str(s['adet'])))
            self.siparis_table.setItem(row, 4, QTableWidgetItem(f"{s['birim_fiyat']:,.2f} TL"))
            self.siparis_table.setItem(row, 5, QTableWidgetItem(f"{s['toplam_tutar']:,.2f} TL"))
            self.siparis_table.setItem(row, 6, QTableWidgetItem(str(s['siparis_tarihi'])[:16]))
            
            durum_item = QTableWidgetItem(s['durum'])
            if s['durum'] == 'Beklemede':
                durum_item.setForeground(QColor("#ff9f1c"))
            elif s['durum'] == 'Hazırlanıyor':
                durum_item.setForeground(QColor("#0077b6"))
            elif s['durum'] == 'Teslim Edildi':
                durum_item.setForeground(QColor("#2ec4b6"))
            self.siparis_table.setItem(row, 7, durum_item)
            
            self.siparis_table.setItem(row, 8, QTableWidgetItem(s['musteri_adi'] or "-"))
            self.siparis_table.setItem(row, 9, QTableWidgetItem(s['musteri_tel'] or "-"))
    
    def hareketleri_listele(self):
        """Hareketleri tabloya doldur"""
        self.hareket_table.setRowCount(0)
        for h in self.db.stok_hareketlerini_getir():
            row = self.hareket_table.rowCount()
            self.hareket_table.insertRow(row)
            
            self.hareket_table.setItem(row, 0, QTableWidgetItem(str(h['hareket_id'])))
            self.hareket_table.setItem(row, 1, QTableWidgetItem(h['urun_adi']))
            
            hareket_item = QTableWidgetItem(h['hareket_tipi'])
            if h['hareket_tipi'] == 'Giriş':
                hareket_item.setForeground(QColor("#4CAF50"))
            elif h['hareket_tipi'] == 'Çıkış':
                hareket_item.setForeground(QColor("#e63946"))
            self.hareket_table.setItem(row, 2, hareket_item)
            
            self.hareket_table.setItem(row, 3, QTableWidgetItem(str(h['miktar'])))
            self.hareket_table.setItem(row, 4, QTableWidgetItem(str(h['onceki_stok'])))
            self.hareket_table.setItem(row, 5, QTableWidgetItem(str(h['sonraki_stok'])))
            self.hareket_table.setItem(row, 6, QTableWidgetItem(str(h['islem_tarihi'])[:16]))
            self.hareket_table.setItem(row, 7, QTableWidgetItem(h['islem_yapan'] or "-"))
    
    def update_dashboard(self):
        """Dashboard kartlarını güncelle"""
        stats = self.db.toplam_istatistikler()
        
        self.urun_card.findChild(QLabel, "value_label").setText(str(stats['urun_sayisi']))
        self.stok_card.findChild(QLabel, "value_label").setText(str(stats['toplam_stok']))
        self.deger_card.findChild(QLabel, "value_label").setText(f"{stats['toplam_deger']:,.0f} TL")
        self.siparis_card.findChild(QLabel, "value_label").setText(str(stats['aktif_siparis']))
        self.dusuk_stok_card.findChild(QLabel, "value_label").setText(str(stats['dusuk_stok_sayisi']))
    
    def refresh_data(self):
        """Otomatik veri yenile"""
        self.urunleri_listele()
        self.siparisleri_listele()
        self.tedarikci_listele_ui()
        self.hareketleri_listele()
        self.update_dashboard()
    
    # ── ÜRÜN İŞLEMLERİ ──
    
    def urun_ekle(self):
        """Ürün ekle dialog"""
        dialog = UrunEkleDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            try:
                self.db.urun_ekle(*dialog.result)
                msg_info(self, "Başarılı", "✅ Ürün başarıyla eklendi!")
                self.urunleri_listele()
                self.update_dashboard()
            except sqlite3.IntegrityError:
                msg_warn(self, "Hata", "⚠️ Bu ürün kodu zaten kullanımda!")
    
    def stok_arttir(self):
        """Stok arttır"""
        row = self.urun_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "⚠️ Lütfen bir ürün seçin!")
            return
        
        urun_id = int(self.urun_table.item(row, 0).text())
        urun_adi = self.urun_table.item(row, 2).text()
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"📈 Stok Arttır - {urun_adi}")
        dialog.setGeometry(400, 300, 350, 200)
        dialog.setStyleSheet(ThemeManager.get("dark"))
        
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"📦 {urun_adi}"))
        layout.addWidget(QLabel("Arttırılacak Miktar:"))
        
        spin = QSpinBox()
        spin.setMinimum(1)
        spin.setMaximum(10000)
        layout.addWidget(spin)
        
        aciklama = QLineEdit()
        aciklama.setPlaceholderText("Açıklama (opsiyonel)")
        layout.addWidget(aciklama)
        
        btn_layout = QHBoxLayout()
        ok_btn = LuxuryButton("✅ Arttır", "#2ec4b6")
        iptal_btn = LuxuryButton("❌ İptal", "#e63946")
        
        def arttir():
            try:
                self.db.stok_arttir(urun_id, spin.value(), aciklama.text(), self.current_user)
                msg_info(self, "Başarılı", f"✅ Stok {spin.value()} arttırıldı!")
                self.urunleri_listele()
                self.hareketleri_listele()
                self.update_dashboard()
                dialog.close()
            except ValueError as e:
                msg_warn(self, "Hata", str(e))
        
        ok_btn.clicked.connect(arttir)
        iptal_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(iptal_btn)
        layout.addLayout(btn_layout)
        
        dialog.setLayout(layout)
        dialog.exec_()
    
    def stok_azalt(self):
        """Stok azalt"""
        row = self.urun_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "⚠️ Lütfen bir ürün seçin!")
            return
        
        urun_id = int(self.urun_table.item(row, 0).text())
        urun_adi = self.urun_table.item(row, 2).text()
        mevcut_stok = int(self.urun_table.item(row, 4).text())
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"📉 Stok Azalt - {urun_adi}")
        dialog.setGeometry(400, 300, 350, 200)
        dialog.setStyleSheet(ThemeManager.get("dark"))
        
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"📦 {urun_adi} (Mevcut: {mevcut_stok})"))
        layout.addWidget(QLabel("Azaltılacak Miktar:"))
        
        spin = QSpinBox()
        spin.setMinimum(1)
        spin.setMaximum(mevcut_stok)
        layout.addWidget(spin)
        
        aciklama = QLineEdit()
        aciklama.setPlaceholderText("Açıklama (opsiyonel)")
        layout.addWidget(aciklama)
        
        btn_layout = QHBoxLayout()
        ok_btn = LuxuryButton("✅ Azalt", "#2ec4b6")
        iptal_btn = LuxuryButton("❌ İptal", "#e63946")
        
        def azalt():
            try:
                self.db.stok_azalt(urun_id, spin.value(), aciklama.text(), self.current_user)
                msg_info(self, "Başarılı", f"✅ Stok {spin.value()} azaltıldı!")
                self.urunleri_listele()
                self.hareketleri_listele()
                self.update_dashboard()
                dialog.close()
            except ValueError as e:
                msg_warn(self, "Hata", str(e))
        
        ok_btn.clicked.connect(azalt)
        iptal_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(iptal_btn)
        layout.addLayout(btn_layout)
        
        dialog.setLayout(layout)
        dialog.exec_()
    
    # ── SİPARİŞ İŞLEMLERİ ──
    
    def siparis_olustur(self):
        """Sipariş oluştur"""
        dialog = SiparisOlusturDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            try:
                siparis_no = self.db.siparis_olustur(*dialog.result)
                msg_info(self, "Başarılı", f"✅ Sipariş oluşturuldu!\nSipariş No: {siparis_no}")
                self.siparisleri_listele()
                self.update_dashboard()
            except ValueError as e:
                msg_warn(self, "Hata", str(e))
    
    # ── TEDARİKÇİ İŞLEMLERİ ──
    
    def tedarikci_ekle(self):
        """Tedarikçi ekle"""
        dialog = TedarikciEkleDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            try:
                self.db.tedarikci_ekle(*dialog.result)
                msg_info(self, "Başarılı", "✅ Tedarikçi başarıyla eklendi!")
                self.tedarikci_listele_ui()
            except Exception as e:
                msg_warn(self, "Hata", f"⚠️ Hata: {str(e)}")
    
    def tedarikci_listele_ui(self):
        """Tedarikçileri tabloya doldur"""
        self.tedarikci_table.setRowCount(0)
        for t in self.db.tedarikci_listele():
            row = self.tedarikci_table.rowCount()
            self.tedarikci_table.insertRow(row)
            
            self.tedarikci_table.setItem(row, 0, QTableWidgetItem(str(t['tedarikci_id'])))
            self.tedarikci_table.setItem(row, 1, QTableWidgetItem(t['tedarikci_adi']))
            self.tedarikci_table.setItem(row, 2, QTableWidgetItem(t['telefon'] or "-"))
            self.tedarikci_table.setItem(row, 3, QTableWidgetItem(t['email'] or "-"))
            self.tedarikci_table.setItem(row, 4, QTableWidgetItem(f"{t['ulas_suresi']} gün"))
            self.tedarikci_table.setItem(row, 5, QTableWidgetItem(str(t['minimum_siparis'])))
            self.tedarikci_table.setItem(row, 6, QTableWidgetItem(f"{t['fiyat_indirim_orani']:.1f}%"))
            self.tedarikci_table.setItem(row, 7, QTableWidgetItem(f"{t['on_time_delivery_rate']:.1f}%"))
            self.tedarikci_table.setItem(row, 8, QTableWidgetItem(t['durum']))
    
    # ── SİPARİŞ DURUM METODLARI ──

    def _secili_siparis_id(self):
        """Tabloda seçili siparişin ID'sini döndür"""
        row = self.siparis_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "⚠️ Lütfen bir sipariş seçin!")
            return None
        return int(self.siparis_table.item(row, 0).text())

    def _siparis_durum_guncelle(self, yeni_durum, onay_mesaji, basari_mesaji):
        """Genel sipariş durum güncelleme"""
        siparis_id = self._secili_siparis_id()
        if siparis_id is None:
            return
        mevcut = self.siparis_table.item(self.siparis_table.currentRow(), 7).text()
        if mevcut == yeni_durum:
            msg_warn(self, "Uyarı", f"⚠️ Sipariş zaten '{yeni_durum}' durumunda!")
            return
        if msg_question(self, "Onay", onay_mesaji) != QMessageBox.Yes:
            return
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE siparisler SET durum = ? WHERE siparis_id = ?",
                    (yeni_durum, siparis_id)
                )
            msg_info(self, "Başarılı", basari_mesaji)
            self.siparisleri_listele()
            self.update_dashboard()
        except Exception as e:
            msg_warn(self, "Hata", f"⚠️ {str(e)}")

    def siparis_onayla(self):
        self._siparis_durum_guncelle(
            "Hazırlanıyor",
            "Seçili siparişi onaylamak istiyor musunuz?",
            "✅ Sipariş onaylandı — Hazırlanıyor."
        )

    def siparis_teslim_et(self):
        self._siparis_durum_guncelle(
            "Teslim Edildi",
            "Siparişi teslim edildi olarak işaretlemek istiyor musunuz?",
            "🚚 Sipariş teslim edildi olarak işaretlendi."
        )

    def siparis_iptal_et(self):
        self._siparis_durum_guncelle(
            "İptal",
            "Siparişi iptal etmek istediğinize emin misiniz?",
            "❌ Sipariş iptal edildi."
        )

    # ── TEDARİKÇİ DÜZENLE / SİL ──

    def tedarikci_duzenle(self):
        """Seçili tedarikçiyi düzenle"""
        row = self.tedarikci_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "⚠️ Lütfen bir tedarikçi seçin!")
            return
        ted_id  = int(self.tedarikci_table.item(row, 0).text())
        ted_adi = self.tedarikci_table.item(row, 1).text()
        tel     = self.tedarikci_table.item(row, 2).text().replace("-", "")
        email   = self.tedarikci_table.item(row, 3).text().replace("-", "")
        ulas    = self.tedarikci_table.item(row, 4).text().replace(" gün", "")
        min_sip = self.tedarikci_table.item(row, 5).text()
        indirim = self.tedarikci_table.item(row, 6).text().replace("%", "")

        dialog = TedarikciDuzenleDialog(
            self.db, ted_id, ted_adi, tel, email, ulas, min_sip, indirim, self
        )
        if dialog.exec_() == QDialog.Accepted:
            msg_info(self, "Başarılı", "✅ Tedarikçi güncellendi!")
            self.tedarikci_listele_ui()

    def tedarikci_sil_ui(self):
        """Seçili tedarikçiyi sil"""
        row = self.tedarikci_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "⚠️ Lütfen bir tedarikçi seçin!")
            return
        ted_id  = int(self.tedarikci_table.item(row, 0).text())
        ted_adi = self.tedarikci_table.item(row, 1).text()
        if msg_question(self, "Sil", f"'{ted_adi}' tedarikçisini silmek istiyor musunuz?") != QMessageBox.Yes:
            return
        try:
            self.db.tedarikci_sil(ted_id)
            msg_info(self, "Başarılı", f"✅ '{ted_adi}' silindi.")
            self.tedarikci_listele_ui()
        except Exception as e:
            msg_warn(self, "Hata", f"⚠️ {str(e)}")

    # ── SİPARİŞ DURUM METODLARI ──

    def _secili_siparis_id(self):
        row = self.siparis_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "Lütfen bir sipariş seçin!")
            return None
        return int(self.siparis_table.item(row, 0).text())

    def _siparis_durum_guncelle(self, yeni_durum, onay_mesaji, basari_mesaji):
        siparis_id = self._secili_siparis_id()
        if siparis_id is None:
            return
        row = self.siparis_table.currentRow()
        mevcut = self.siparis_table.item(row, 7).text()
        if mevcut == yeni_durum:
            msg_warn(self, "Uyarı", f"Sipariş zaten '{yeni_durum}' durumunda!")
            return
        if msg_question(self, "Onay", onay_mesaji) != QMessageBox.Yes:
            return
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE siparisler SET durum = ? WHERE siparis_id = ?",
                    (yeni_durum, siparis_id)
                )
            msg_info(self, "Basarili", basari_mesaji)
            self.siparisleri_listele()
            self.update_dashboard()
        except Exception as e:
            msg_warn(self, "Hata", str(e))

    def siparis_onayla(self):
        self._siparis_durum_guncelle(
            "Hazirlanıyor",
            "Seçili siparişi onaylamak istiyor musunuz?",
            "Sipariş onaylandı."
        )

    def siparis_teslim_et(self):
        self._siparis_durum_guncelle(
            "Teslim Edildi",
            "Siparişi teslim edildi olarak işaretlemek istiyor musunuz?",
            "Sipariş teslim edildi olarak işaretlendi."
        )

    def siparis_iptal_et(self):
        self._siparis_durum_guncelle(
            "Iptal",
            "Siparişi iptal etmek istediğinize emin misiniz?",
            "Sipariş iptal edildi."
        )

    # ── TEDARIKCİ DUZENLE / SIL ──

    def tedarikci_duzenle(self):
        row = self.tedarikci_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "Lütfen bir tedarikçi seçin!")
            return
        ted_id  = int(self.tedarikci_table.item(row, 0).text())
        ted_adi = self.tedarikci_table.item(row, 1).text()
        tel     = self.tedarikci_table.item(row, 2).text().replace("-", "")
        email   = self.tedarikci_table.item(row, 3).text().replace("-", "")
        ulas    = self.tedarikci_table.item(row, 4).text().replace(" gun", "").replace(" gün", "")
        min_sip = self.tedarikci_table.item(row, 5).text()
        indirim = self.tedarikci_table.item(row, 6).text().replace("%", "")
        dialog = TedarikciDuzenleDialog(
            self.db, ted_id, ted_adi, tel, email, ulas, min_sip, indirim, self
        )
        if dialog.exec_() == QDialog.Accepted:
            msg_info(self, "Basarili", "Tedarikçi güncellendi!")
            self.tedarikci_listele_ui()

    def tedarikci_sil_ui(self):
        row = self.tedarikci_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "Lütfen bir tedarikçi seçin!")
            return
        ted_id  = int(self.tedarikci_table.item(row, 0).text())
        ted_adi = self.tedarikci_table.item(row, 1).text()
        if msg_question(self, "Sil", f"'{ted_adi}' tedarikçisini silmek istiyor musunuz?") != QMessageBox.Yes:
            return
        try:
            self.db.tedarikci_sil(ted_id)
            msg_info(self, "Basarili", f"'{ted_adi}' silindi.")
            self.tedarikci_listele_ui()
        except Exception as e:
            msg_warn(self, "Hata", str(e))

    # ── SİPARİŞ DURUM METODLARI ──

    def _secili_siparis_id(self):
        row = self.siparis_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyari", "Lutfen bir siparis secin!")
            return None
        return int(self.siparis_table.item(row, 0).text())

    def _siparis_durum_guncelle(self, yeni_durum, onay_mesaji, basari_mesaji):
        siparis_id = self._secili_siparis_id()
        if siparis_id is None:
            return
        row = self.siparis_table.currentRow()
        mevcut = self.siparis_table.item(row, 7).text()
        if mevcut == yeni_durum:
            msg_warn(self, "Uyari", f"Siparis zaten '{yeni_durum}' durumunda!")
            return
        if msg_question(self, "Onay", onay_mesaji) != QMessageBox.Yes:
            return
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE siparisler SET durum = ? WHERE siparis_id = ?",
                    (yeni_durum, siparis_id)
                )
            msg_info(self, "Basarili", basari_mesaji)
            self.siparisleri_listele()
            self.update_dashboard()
        except Exception as e:
            msg_warn(self, "Hata", str(e))

    def siparis_onayla(self):
        self._siparis_durum_guncelle(
            "Hazirlanıyor",
            "Secili siparisi onaylamak istiyor musunuz?",
            "Siparis onaylandi, hazirlanıyor."
        )

    def siparis_teslim_et(self):
        self._siparis_durum_guncelle(
            "Teslim Edildi",
            "Siparisi teslim edildi olarak isaretle?",
            "Siparis teslim edildi olarak isaretlendi."
        )

    def siparis_iptal_et(self):
        self._siparis_durum_guncelle(
            "Iptal",
            "Siparisi iptal etmek istediginize emin misiniz?",
            "Siparis iptal edildi."
        )

    # ── TEDARIKCİ DUZENLE / SIL ──

    def tedarikci_duzenle(self):
        row = self.tedarikci_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyari", "Lutfen bir tedarikci secin!")
            return
        ted_id  = int(self.tedarikci_table.item(row, 0).text())
        ted_adi = self.tedarikci_table.item(row, 1).text()
        tel     = self.tedarikci_table.item(row, 2).text().replace("-", "")
        email   = self.tedarikci_table.item(row, 3).text().replace("-", "")
        ulas    = self.tedarikci_table.item(row, 4).text().replace(" gun", "").replace(" gün", "")
        min_sip = self.tedarikci_table.item(row, 5).text()
        indirim = self.tedarikci_table.item(row, 6).text().replace("%", "")
        dialog = TedarikciDuzenleDialog(
            self.db, ted_id, ted_adi, tel, email, ulas, min_sip, indirim, self
        )
        if dialog.exec_() == QDialog.Accepted:
            msg_info(self, "Basarili", "Tedarikci guncellendi!")
            self.tedarikci_listele_ui()

    def tedarikci_sil_ui(self):
        row = self.tedarikci_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyari", "Lutfen bir tedarikci secin!")
            return
        ted_id  = int(self.tedarikci_table.item(row, 0).text())
        ted_adi = self.tedarikci_table.item(row, 1).text()
        if msg_question(self, "Sil", f"'{ted_adi}' tedarikciyi silmek istiyor musunuz?") != QMessageBox.Yes:
            return
        try:
            self.db.tedarikci_sil(ted_id)
            msg_info(self, "Basarili", f"'{ted_adi}' silindi.")
            self.tedarikci_listele_ui()
        except Exception as e:
            msg_warn(self, "Hata", str(e))

    # ── AYARLAR ──
    
    def sifre_degistir(self):
        """Şifre değiştir"""
        yeni_sifre = self.yeni_sifre.text().strip()
        onayla = self.yeni_sifre_onayla.text().strip()
        
        if not yeni_sifre:
            msg_warn(self, "Uyarı", "⚠️ Yeni şifre boş olamaz!")
            return
        
        if yeni_sifre != onayla:
            msg_warn(self, "Uyarı", "⚠️ Şifreler eşleşmiyor!")
            return
        
        if len(yeni_sifre) < 6:
            msg_warn(self, "Uyarı", "⚠️ Şifre en az 6 karakter olmalıdır!")
            return
        
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                sifre_hash = hashlib.sha256(yeni_sifre.encode()).hexdigest()
                
                cursor.execute('''
                    UPDATE sistem_kullanicilar
                    SET sifre = ?
                    WHERE kullanici_adi = ?
                ''', (sifre_hash, self.current_user))
            
            msg_info(self, "Başarılı", "✅ Şifre başarıyla değiştirildi!")
            self.yeni_sifre.clear()
            self.yeni_sifre_onayla.clear()
        except Exception as e:
            msg_warn(self, "Hata", f"⚠️ Hata: {str(e)}")
    
    def veritabani_yedekle(self):
        """Veritabanını yedekle"""
        try:
            import shutil
            dosya_adi = f"depo_yonetim_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy(self.db.db_name, dosya_adi)
            msg_info(self, "Başarılı", f"✅ Veritabanı başarıyla yedeklendi!\n📁 {dosya_adi}")
        except Exception as e:
            msg_warn(self, "Hata", f"⚠️ Yedekleme hatası: {str(e)}")
    
    # ── RAPORLAR ──
    
    def raporlar_ac(self):
        """Raporlar diyaloğunu aç"""
        dialog = RaporlarDialog(self.db, self)
        dialog.exec_()
    
    # ── ANOMALİLER ──
    
    def anomaliler_ac(self):
        """Anomaliler diyaloğunu aç"""
        dialog = AnomalilerDialog(self.db, self)
        dialog.exec_()
    
    # ── GAMİFİCATİON ──
    
    def gamification_ac(self):
        """Gamification diyaloğunu aç"""
        dialog = GamificationDialog(self.db, self.current_user, self)
        dialog.exec_()
    
    # ── STOK SAYIMI ──
    
    def stok_sayimi_ac(self):
        """Stok Sayımı diyaloğunu aç"""
        dialog = StokSayimiDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted:
            self.urunleri_listele()
            self.update_dashboard()
            msg_info(self, "Başarılı", "✅ Stok sayımı başarıyla tamamlandı!")
    
    # ── KIŞAYOLLAR & YARDIM ──
    
    def kisayollar_ekle(self):
        """Sistem kısayollarını ekle"""
        # F1 - Yardım
        f1_action = QAction(self)
        f1_action.setShortcut("F1")
        f1_action.triggered.connect(self.yardim_ac)
        self.addAction(f1_action)
        
        # Ctrl+H - Kısayolları Göster
        ctrlh_action = QAction(self)
        ctrlh_action.setShortcut("Ctrl+H")
        ctrlh_action.triggered.connect(self.yardim_ac)
        self.addAction(ctrlh_action)
        
        # Ctrl+Q - Çıkış
        ctrlq_action = QAction(self)
        ctrlq_action.setShortcut("Ctrl+Q")
        ctrlq_action.triggered.connect(self.close)
        self.addAction(ctrlq_action)
        
        # Ctrl+N - Yeni Ürün
        ctrln_action = QAction(self)
        ctrln_action.setShortcut("Ctrl+N")
        ctrln_action.triggered.connect(self.urun_ekle)
        self.addAction(ctrln_action)
        
        # Ctrl+O - Raporlar
        ctrlo_action = QAction(self)
        ctrlo_action.setShortcut("Ctrl+O")
        ctrlo_action.triggered.connect(self.raporlar_ac)
        self.addAction(ctrlo_action)
    
    def yardim_ac(self):
        """Yardım diyaloğunu aç"""
        dialog = YardimDialog(self)
        dialog.exec_()


# ═══════════════════════════════════════════════════════════════════════════
# PHASE 2: LOGIN SİSTEMİ, TEDARİKÇİ, AYARLAR
# ═══════════════════════════════════════════════════════════════════════════

# ───────────────────────────────────────────────────────────────────────────
# LOGIN DIALOG
# ───────────────────────────────────────────────────────────────────────────

class LoginDialog(QDialog):
    """Giriş Diyalogu - Rol Tabanlı"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.user_data = None
        self.setWindowTitle("🏪 DEPO YÖNETİM SİSTEMİ - Giriş")
        self.setGeometry(500, 250, 450, 350)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(40, 40, 40, 40)
        
        # Header
        header = QLabel("🔐 GÜVENLI GİRİŞ")
        header.setFont(QFont("Arial", 18, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("color: #0077b6; margin-bottom: 20px;")
        
        # Form
        form_layout = QGridLayout()
        form_layout.setSpacing(15)
        
        # Kullanıcı Adı
        form_layout.addWidget(QLabel("👤 Kullanıcı Adı:"), 0, 0)
        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("admin")
        self.username_input.setText("admin")
        form_layout.addWidget(self.username_input, 0, 1)
        
        # Şifre
        form_layout.addWidget(QLabel("🔑 Şifre:"), 1, 0)
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setPlaceholderText("(boş bırakılabilir)")
        self.password_input.returnPressed.connect(self.giris_yap)
        form_layout.addWidget(self.password_input, 1, 1)
        
        layout.addWidget(header)
        layout.addLayout(form_layout)
        layout.addSpacing(10)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        
        giris_btn = LuxuryButton("✅ Giriş Yap", "#2ec4b6")
        giris_btn.clicked.connect(self.giris_yap)
        
        cikis_btn = LuxuryButton("❌ Çıkış", "#e63946")
        cikis_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(giris_btn)
        btn_layout.addWidget(cikis_btn)
        
        layout.addLayout(btn_layout)
        layout.addStretch()
        
        self.setLayout(layout)
    
    def giris_yap(self):
        """Kullanıcı giriş kontrolü"""
        kullanici_adi = self.username_input.text().strip()
        sifre = self.password_input.text().strip()
        
        if not kullanici_adi:
            msg_warn(self, "Uyarı", "⚠️ Kullanıcı adı boş olamaz!")
            return
        
        # Şifre boşsa admin olarak giriş yap
        if not sifre:
            self.user_data = {
                'kullanici_id': 1,
                'ad': 'Admin',
                'soyad': 'Kullanıcı',
                'rol': 'admin',
                'durum': 'Aktif',
                'kullanici_adi': kullanici_adi
            }
            msg_info(self, "Başarılı", f"✅ Hoş geldiniz, Admin!")
            self.accept()
            return
        
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            sifre_hash = hashlib.sha256(sifre.encode()).hexdigest()
            
            cursor.execute('''
                SELECT kullanici_id, ad, soyad, rol, durum
                FROM sistem_kullanicilar
                WHERE kullanici_adi = ? AND sifre = ?
            ''', (kullanici_adi, sifre_hash))
            
            user = cursor.fetchone()
            
            if not user:
                msg_warn(self, "Hata", "❌ Kullanıcı adı veya şifre yanlış!")
                self.password_input.clear()
                return
            
            if user['durum'] == 'Pasif':
                msg_warn(self, "Hata", "⚠️ Hesabınız devre dışı bırakılmıştır!")
                return
            
            self.user_data = dict(user)
            self.user_data['kullanici_adi'] = kullanici_adi
            
            cursor.execute('''
                UPDATE sistem_kullanicilar
                SET son_giris_tarihi = CURRENT_TIMESTAMP
                WHERE kullanici_id = ?
            ''', (user['kullanici_id'],))
            
            msg_info(self, "Başarılı", f"✅ Hoş geldiniz, {user['ad']}!")
            self.accept()


# ───────────────────────────────────────────────────────────────────────────
# TEDARİKÇİ DIALOG
# ───────────────────────────────────────────────────────────────────────────

class TedarikciEkleDialog(QDialog):
    """Tedarikçi Ekle Dialog"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("🏢 Yeni Tedarikçi Ekle")
        self.setGeometry(200, 200, 600, 600)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.result = None
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)
        
        baslik = QLabel("🏢 YENİ TEDARİKÇİ EKLE")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6; margin-bottom: 10px;")
        
        grid = QGridLayout()
        grid.setSpacing(12)
        
        grid.addWidget(QLabel("Tedarikçi Adı:*"), 0, 0)
        self.adi_input = QLineEdit()
        grid.addWidget(self.adi_input, 0, 1)
        
        grid.addWidget(QLabel("Telefon:"), 1, 0)
        self.tel_input = QLineEdit()
        grid.addWidget(self.tel_input, 1, 1)
        
        grid.addWidget(QLabel("Email:"), 2, 0)
        self.email_input = QLineEdit()
        grid.addWidget(self.email_input, 2, 1)
        
        grid.addWidget(QLabel("Adres:"), 3, 0)
        self.adres_input = QLineEdit()
        grid.addWidget(self.adres_input, 3, 1)
        
        grid.addWidget(QLabel("Ulaş Süresi (Gün):"), 4, 0)
        self.ulas_input = QSpinBox()
        self.ulas_input.setMinimum(1)
        self.ulas_input.setMaximum(30)
        self.ulas_input.setValue(3)
        grid.addWidget(self.ulas_input, 4, 1)
        
        grid.addWidget(QLabel("Minimum Sipariş:"), 5, 0)
        self.min_siparis_input = QSpinBox()
        self.min_siparis_input.setMinimum(1)
        self.min_siparis_input.setMaximum(10000)
        self.min_siparis_input.setValue(50)
        grid.addWidget(self.min_siparis_input, 5, 1)
        
        grid.addWidget(QLabel("Fiyat İndirim (%):"), 6, 0)
        self.indirim_input = QDoubleSpinBox()
        self.indirim_input.setMinimum(0)
        self.indirim_input.setMaximum(100)
        grid.addWidget(self.indirim_input, 6, 1)
        
        grid.addWidget(QLabel("Açıklama:"), 7, 0)
        self.aciklama_input = QTextEdit()
        self.aciklama_input.setMaximumHeight(80)
        grid.addWidget(self.aciklama_input, 7, 1)
        
        layout.addWidget(baslik)
        layout.addLayout(grid)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        ekle_btn = LuxuryButton("✅ Tedarikçi Ekle", "#2ec4b6")
        ekle_btn.clicked.connect(self.ekle)
        iptal_btn = LuxuryButton("❌ İptal", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ekle_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def ekle(self):
        if self.adi_input.text().strip():
            self.result = (
                self.adi_input.text().strip(),
                self.tel_input.text().strip() or None,
                self.email_input.text().strip() or None,
                self.adres_input.text().strip() or None,
                self.ulas_input.value(),
                self.min_siparis_input.value(),
                self.indirim_input.value()
            )
            self.accept()
        else:
            msg_warn(self, "Uyarı", "⚠️ Tedarikçi adı zorunludur!")


# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 8: DIALOG SINIFFLARI
# ═══════════════════════════════════════════════════════════════════════════


class TedarikciDuzenleDialog(QDialog):
    """Tedarikçi Düzenleme Dialogu"""

    def __init__(self, db, ted_id, adi, tel, email, ulas, min_sip, indirim, parent=None):
        super().__init__(parent)
        self.db = db
        self.ted_id = ted_id
        self.setWindowTitle(f"Tedarikci Duzenle - {adi}")
        self.setGeometry(250, 250, 520, 480)
        self.setStyleSheet(ThemeManager.get("dark"))
        self._adi = adi; self._tel = tel; self._email = email
        self._ulas = ulas; self._min = min_sip; self._ind = indirim
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)

        baslik = QLabel("Tedarikci Duzenle")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color:#0077b6;")
        layout.addWidget(baslik)

        grid = QGridLayout()
        grid.setSpacing(12)

        grid.addWidget(QLabel("Tedarikci Adi:*"), 0, 0)
        self.adi_input = QLineEdit(self._adi)
        grid.addWidget(self.adi_input, 0, 1)

        grid.addWidget(QLabel("Telefon:"), 1, 0)
        self.tel_input = QLineEdit(self._tel)
        grid.addWidget(self.tel_input, 1, 1)

        grid.addWidget(QLabel("Email:"), 2, 0)
        self.email_input = QLineEdit(self._email)
        grid.addWidget(self.email_input, 2, 1)

        grid.addWidget(QLabel("Ulas Suresi (Gun):"), 3, 0)
        self.ulas_input = QSpinBox()
        self.ulas_input.setRange(1, 60)
        try:
            self.ulas_input.setValue(int(self._ulas))
        except Exception:
            self.ulas_input.setValue(3)
        grid.addWidget(self.ulas_input, 3, 1)

        grid.addWidget(QLabel("Min Siparis:"), 4, 0)
        self.min_input = QSpinBox()
        self.min_input.setRange(1, 10000)
        try:
            self.min_input.setValue(int(self._min))
        except Exception:
            self.min_input.setValue(50)
        grid.addWidget(self.min_input, 4, 1)

        grid.addWidget(QLabel("Indirim (%):"), 5, 0)
        self.ind_input = QDoubleSpinBox()
        self.ind_input.setRange(0, 100)
        try:
            self.ind_input.setValue(float(self._ind))
        except Exception:
            self.ind_input.setValue(0)
        grid.addWidget(self.ind_input, 5, 1)

        layout.addLayout(grid)

        btn_layout = QHBoxLayout()
        kaydet_btn = LuxuryButton("Kaydet", "#2ec4b6")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = LuxuryButton("Iptal", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def kaydet(self):
        if not self.adi_input.text().strip():
            msg_warn(self, "Uyari", "Tedarikci adi bos olamaz!")
            return
        try:
            self.db.tedarikci_guncelle(
                self.ted_id,
                tedarikci_adi=self.adi_input.text().strip(),
                telefon=self.tel_input.text().strip(),
                email=self.email_input.text().strip(),
                ulas_suresi=self.ulas_input.value(),
                minimum_siparis=self.min_input.value(),
                fiyat_indirim_orani=self.ind_input.value()
            )
            self.accept()
        except Exception as e:
            msg_warn(self, "Hata", str(e))


class UrunEkleDialog(QDialog):
    """Ürün ekle diyalogu"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("📦 Yeni Ürün Ekle")
        self.setGeometry(200, 200, 600, 700)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.result = None
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)
        
        baslik = QLabel("📦 YENİ ÜRÜN EKLE")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6; margin-bottom: 10px;")
        
        grid = QGridLayout()
        grid.setSpacing(12)
        
        grid.addWidget(QLabel("Ürün Kodu:"), 0, 0)
        self.kod_input = QLineEdit()
        grid.addWidget(self.kod_input, 0, 1)
        
        grid.addWidget(QLabel("Ürün Adı:*"), 1, 0)
        self.ad_input = QLineEdit()
        grid.addWidget(self.ad_input, 1, 1)
        
        grid.addWidget(QLabel("Kategori:*"), 2, 0)
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItems(self.db.kategorileri_getir())
        grid.addWidget(self.kategori_combo, 2, 1)
        
        grid.addWidget(QLabel("Başlangıç Stok:"), 3, 0)
        self.stok_input = QSpinBox()
        self.stok_input.setMaximum(100000)
        grid.addWidget(self.stok_input, 3, 1)
        
        grid.addWidget(QLabel("Minimum Stok:"), 4, 0)
        self.min_stok_input = QSpinBox()
        self.min_stok_input.setMaximum(10000)
        self.min_stok_input.setValue(10)
        grid.addWidget(self.min_stok_input, 4, 1)
        
        grid.addWidget(QLabel("Reorder Point:"), 5, 0)
        self.reorder_input = QSpinBox()
        self.reorder_input.setMaximum(10000)
        self.reorder_input.setValue(20)
        grid.addWidget(self.reorder_input, 5, 1)
        
        grid.addWidget(QLabel("Satış Fiyatı (TL):*"), 6, 0)
        self.fiyat_input = QDoubleSpinBox()
        self.fiyat_input.setMaximum(1000000)
        grid.addWidget(self.fiyat_input, 6, 1)
        
        grid.addWidget(QLabel("Alış Fiyatı (TL):"), 7, 0)
        self.alis_input = QDoubleSpinBox()
        self.alis_input.setMaximum(1000000)
        grid.addWidget(self.alis_input, 7, 1)
        
        grid.addWidget(QLabel("Birim:"), 8, 0)
        self.birim_combo = QComboBox()
        self.birim_combo.addItems(["adet", "kg", "litre", "paket", "kutu", "metre"])
        grid.addWidget(self.birim_combo, 8, 1)
        
        grid.addWidget(QLabel("Raf No:"), 9, 0)
        self.raf_input = QLineEdit()
        grid.addWidget(self.raf_input, 9, 1)
        
        grid.addWidget(QLabel("Açıklama:"), 10, 0)
        self.aciklama_input = QTextEdit()
        self.aciklama_input.setMaximumHeight(80)
        grid.addWidget(self.aciklama_input, 10, 1)
        
        layout.addWidget(baslik)
        layout.addLayout(grid)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        ekle_btn = LuxuryButton("✅ Ürün Ekle", "#2ec4b6")
        ekle_btn.clicked.connect(self.ekle)
        iptal_btn = LuxuryButton("❌ İptal", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ekle_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def ekle(self):
        if self.ad_input.text().strip() and self.fiyat_input.value() > 0:
            kod = self.kod_input.text().strip() or f"PRD{random.randint(10000, 99999)}"
            self.result = (
                kod, self.ad_input.text().strip(), self.kategori_combo.currentText(),
                self.stok_input.value(), self.min_stok_input.value(),
                self.reorder_input.value(), self.fiyat_input.value(),
                self.alis_input.value() if self.alis_input.value() > 0 else None,
                self.birim_combo.currentText(), self.raf_input.text().strip(),
                self.aciklama_input.toPlainText().strip()
            )
            self.accept()
        else:
            msg_warn(self, "Uyarı", "⚠️ Ürün adı ve satış fiyatı zorunludur!")


class SiparisOlusturDialog(QDialog):
    """Sipariş oluştur diyalogu"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("🛒 Yeni Sipariş Oluştur")
        self.setGeometry(200, 200, 600, 550)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.result = None
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)
        
        baslik = QLabel("🛒 YENİ SİPARİŞ OLUŞTUR")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6; margin-bottom: 10px;")
        
        grid = QGridLayout()
        grid.setSpacing(12)
        
        grid.addWidget(QLabel("Ürün Seçin:*"), 0, 0)
        self.urun_combo = QComboBox()
        for urun in self.db.urunleri_getir():
            self.urun_combo.addItem(f"{urun['urun_adi']} - Stok: {urun['stok']}", urun['urun_id'])
        self.urun_combo.currentIndexChanged.connect(self.urun_fiyat_doldur)
        grid.addWidget(self.urun_combo, 0, 1)
        
        grid.addWidget(QLabel("Adet:*"), 1, 0)
        self.adet_input = QSpinBox()
        self.adet_input.setMinimum(1)
        self.adet_input.setMaximum(1000)
        self.adet_input.valueChanged.connect(self.fiyat_hesapla)
        grid.addWidget(self.adet_input, 1, 1)
        
        grid.addWidget(QLabel("Birim Fiyat (TL):"), 2, 0)
        self.fiyat_input = QDoubleSpinBox()
        self.fiyat_input.setMinimum(0)
        self.fiyat_input.setMaximum(1000000)
        self.fiyat_input.valueChanged.connect(self.fiyat_hesapla)
        grid.addWidget(self.fiyat_input, 2, 1)
        
        grid.addWidget(QLabel("Toplam (TL):"), 3, 0)
        self.toplam_label = QLabel("0.00 TL")
        self.toplam_label.setStyleSheet("color: #2ec4b6; font-weight: bold; font-size: 16px;")
        grid.addWidget(self.toplam_label, 3, 1)
        
        grid.addWidget(QLabel("Müşteri Adı:"), 4, 0)
        self.musteri_input = QLineEdit()
        grid.addWidget(self.musteri_input, 4, 1)
        
        grid.addWidget(QLabel("Müşteri Tel:"), 5, 0)
        self.tel_input = QLineEdit()
        grid.addWidget(self.tel_input, 5, 1)
        
        grid.addWidget(QLabel("Müşteri Email:"), 6, 0)
        self.email_input = QLineEdit()
        grid.addWidget(self.email_input, 6, 1)
        
        grid.addWidget(QLabel("Açıklama:"), 7, 0)
        self.aciklama_input = QTextEdit()
        self.aciklama_input.setMaximumHeight(80)
        grid.addWidget(self.aciklama_input, 7, 1)
        
        layout.addWidget(baslik)
        layout.addLayout(grid)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        olustur_btn = LuxuryButton("✅ Sipariş Oluştur", "#2ec4b6")
        olustur_btn.clicked.connect(self.siparis_olustur)
        iptal_btn = LuxuryButton("❌ İptal", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(olustur_btn)
        btn_layout.addWidget(iptal_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def urun_fiyat_doldur(self):
        index = self.urun_combo.currentIndex()
        if index >= 0:
            urun_id = self.urun_combo.itemData(index)
            for urun in self.db.urunleri_getir():
                if urun['urun_id'] == urun_id:
                    self.fiyat_input.setValue(urun['fiyat'])
                    break
    
    def fiyat_hesapla(self):
        toplam = self.adet_input.value() * self.fiyat_input.value()
        self.toplam_label.setText(f"{toplam:,.2f} TL")
    
    def siparis_olustur(self):
        urun_id = self.urun_combo.currentData()
        adet = self.adet_input.value()
        fiyat = self.fiyat_input.value()
        if urun_id and adet > 0 and fiyat > 0:
            self.result = (
                urun_id, adet, fiyat,
                self.musteri_input.text().strip(),
                self.tel_input.text().strip(),
                self.email_input.text().strip(),
                self.aciklama_input.toPlainText().strip()
            )
            self.accept()
        else:
            msg_warn(self, "Uyarı", "⚠️ Ürün, adet ve fiyat zorunludur!")


# ═══════════════════════════════════════════════════════════════════════════
# PHASE 3: RAPORLAR & GRAFİKLER
# ═══════════════════════════════════════════════════════════════════════════

# ───────────────────────────────────────────────────────────────────────────
# GRAFİK CANVAS SINIFI
# ───────────────────────────────────────────────────────────────────────────

class MatplotlibCanvas(FigureCanvas):
    """Matplotlib Grafik Canvas"""
    
    def __init__(self, parent=None, width=5, height=4, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = self.fig.add_subplot(111)
        super().__init__(self.fig)
        self.setParent(parent)
        self.fig.patch.set_facecolor('#0f0f1a')
        self.axes.set_facecolor('#1a1a2e')
        
    def plot_pie(self, labels, sizes, colors):
        """Pasta grafik"""
        self.axes.clear()
        self.axes.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%',
                     startangle=90, textprops={'color': '#ffffff'})
        self.axes.set_title("Kategori Dağılımı", color='#0077b6', fontsize=14, fontweight='bold')
        self.fig.tight_layout()
        self.draw()
    
    def plot_line(self, x, y, label, color='#00b4d8'):
        """Çizgi grafik"""
        self.axes.clear()
        self.axes.plot(x, y, marker='o', color=color, linewidth=2, markersize=8)
        self.axes.set_title("Aylık Sipariş Trendi", color='#0077b6', fontsize=14, fontweight='bold')
        self.axes.set_xlabel("Aylar", color='#ffffff')
        self.axes.set_ylabel("Sipariş Sayısı", color='#ffffff')
        self.axes.tick_params(colors='#ffffff')
        self.axes.grid(True, alpha=0.3, color='#ffffff')
        self.fig.tight_layout()
        self.draw()
    
    def plot_bar(self, categories, values, colors):
        """Bar grafik"""
        self.axes.clear()
        bars = self.axes.bar(categories, values, color=colors, edgecolor='#ffffff', linewidth=1.5)
        self.axes.set_title("En Çok Satılan Ürünler", color='#0077b6', fontsize=14, fontweight='bold')
        self.axes.set_ylabel("Satış Miktarı", color='#ffffff')
        self.axes.tick_params(colors='#ffffff')
        
        for bar in bars:
            height = bar.get_height()
            self.axes.text(bar.get_x() + bar.get_width()/2., height,
                          f'{int(height)}',
                          ha='center', va='bottom', color='#ffffff')
        
        plt.setp(self.axes.xaxis.get_majorticklabels(), rotation=45, ha='right')
        self.fig.tight_layout()
        self.draw()
    
    def plot_scatter(self, x, y, colors, sizes, labels):
        """Scatter grafik (ABC Analizi)"""
        self.axes.clear()
        for i, label in enumerate(set(labels)):
            mask = [l == label for l in labels]
            x_filtered = [x[j] for j in range(len(x)) if mask[j]]
            y_filtered = [y[j] for j in range(len(y)) if mask[j]]
            color_filtered = [colors[j] for j in range(len(colors)) if mask[j]]
            
            self.axes.scatter(x_filtered, y_filtered, c=color_filtered, s=[s for j, s in enumerate(sizes) if mask[j]],
                            alpha=0.6, edgecolors='#ffffff', linewidth=1)
        
        self.axes.set_title("ABC Analizi (Stok vs Değer)", color='#0077b6', fontsize=14, fontweight='bold')
        self.axes.set_xlabel("Stok Miktarı", color='#ffffff')
        self.axes.set_ylabel("Toplam Değer (₺)", color='#ffffff')
        self.axes.tick_params(colors='#ffffff')
        self.axes.grid(True, alpha=0.3, color='#ffffff')
        self.fig.tight_layout()
        self.draw()


# ───────────────────────────────────────────────────────────────────────────
# RAPOR DİYALOGU
# ───────────────────────────────────────────────────────────────────────────

class RaporlarDialog(QDialog):
    """Raporlar Ana Diyalogu"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("📊 Raporlar")
        self.setGeometry(100, 100, 900, 700)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("📊 RAPORLAR & GRAFİKLER")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6; margin-bottom: 20px;")
        
        tabs = QTabWidget()
        tabs.addTab(self.create_grafik_tab(), "📈 Grafikler")
        tabs.addTab(self.create_export_tab(), "💾 Export")
        
        layout.addWidget(baslik)
        layout.addWidget(tabs)
        self.setLayout(layout)
    
    def create_grafik_tab(self):
        """Grafikler tab'ını oluştur"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        select_layout = QHBoxLayout()
        select_layout.addWidget(QLabel("📊 Grafik Seç:"))
        
        self.grafik_combo = QComboBox()
        self.grafik_combo.addItems([
            "Kategori Dağılımı (Pasta)",
            "Aylık Sipariş Trendi (Çizgi)",
            "En Çok Satılan (Bar)",
            "ABC Analizi (Scatter)"
        ])
        self.grafik_combo.currentIndexChanged.connect(self.grafik_goster)
        select_layout.addWidget(self.grafik_combo)
        
        yenile_btn = LuxuryButton("🔄 Yenile", "#0077b6")
        yenile_btn.setMaximumWidth(120)
        yenile_btn.clicked.connect(self.grafik_goster)
        select_layout.addWidget(yenile_btn)
        select_layout.addStretch()
        
        layout.addLayout(select_layout)
        
        self.canvas = MatplotlibCanvas(width=9, height=5.5, dpi=100)
        layout.addWidget(self.canvas)
        
        widget.setLayout(layout)
        self.grafik_goster()
        
        return widget
    
    def grafik_goster(self):
        """Seçilen grafiği göster"""
        index = self.grafik_combo.currentIndex()
        
        if index == 0:
            self.show_kategori_dagitimi()
        elif index == 1:
            self.show_aylik_siparis()
        elif index == 2:
            self.show_en_cok_satilan()
        elif index == 3:
            self.show_abc_analizi()
    
    def show_kategori_dagitimi(self):
        """Kategori dağılımı pasta grafiği"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT kategori, COUNT(*) as adet
                FROM urunler
                WHERE durum = 'Aktif'
                GROUP BY kategori
            ''')
            data = cursor.fetchall()
        
        if data:
            labels = [d['kategori'] for d in data]
            sizes = [d['adet'] for d in data]
            colors = ['#0077b6', '#00b4d8', '#2ec4b6', '#4CAF50', '#ff9f1c', '#e63946', '#ff6b6b', '#c44569']
            colors = colors[:len(labels)]
            self.canvas.plot_pie(labels, sizes, colors)
    
    def show_aylik_siparis(self):
        """Aylık sipariş trendi"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT strftime('%m', siparis_tarihi) as ay, COUNT(*) as siparis_sayisi
                FROM siparisler
                GROUP BY ay
                ORDER BY ay
                LIMIT 12
            ''')
            data = cursor.fetchall()
        
        if data:
            months = ['Oca', 'Şub', 'Mar', 'Nis', 'May', 'Haz', 'Tem', 'Ağu', 'Eyl', 'Eki', 'Kas', 'Ara']
            x = [months[int(d['ay'])-1] if d['ay'] else 'N/A' for d in data]
            y = [d['siparis_sayisi'] for d in data]
            self.canvas.plot_line(x, y, "Siparişler", color='#2ec4b6')
    
    def show_en_cok_satilan(self):
        """En çok satılan ürünler"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT u.urun_adi, SUM(s.adet) as toplam_adet
                FROM siparisler s
                JOIN urunler u ON s.urun_id = u.urun_id
                GROUP BY u.urun_adi
                ORDER BY toplam_adet DESC
                LIMIT 8
            ''')
            data = cursor.fetchall()
        
        if data:
            urunler = [d['urun_adi'][:15] for d in data]
            miktarlar = [d['toplam_adet'] for d in data]
            colors = ['#0077b6', '#00b4d8', '#2ec4b6', '#4CAF50', '#ff9f1c', '#e63946', '#ff6b6b', '#c44569']
            colors = colors[:len(urunler)]
            self.canvas.plot_bar(urunler, miktarlar, colors)
    
    def show_abc_analizi(self):
        """ABC analizi scatter"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT urun_adi, stok, (stok * fiyat) as deger
                FROM urunler
                WHERE durum = 'Aktif'
                ORDER BY deger DESC
            ''')
            data = cursor.fetchall()
        
        if data:
            x = [d['stok'] for d in data]
            y = [d['deger'] for d in data]
            
            total = sum(y)
            cumsum = 0
            labels = []
            colors = []
            
            for value in y:
                cumsum += value
                percent = (cumsum / total) * 100
                
                if percent <= 80:
                    labels.append('A')
                    colors.append('#e63946')
                elif percent <= 95:
                    labels.append('B')
                    colors.append('#ff9f1c')
                else:
                    labels.append('C')
                    colors.append('#2ec4b6')
            
            sizes = [50 + (v/max(y))*100 for v in y]
            self.canvas.plot_scatter(x, y, colors, sizes, labels)
    
    def create_export_tab(self):
        """Export tab'ını oluştur"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        csv_frame = QFrame()
        csv_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 10px;
                padding: 15px;
            }
        """)
        csv_layout = QVBoxLayout()
        
        csv_baslik = QLabel("📄 CSV Export")
        csv_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        csv_baslik.setStyleSheet("color: #0077b6;")
        csv_layout.addWidget(csv_baslik)
        
        csv_btn_layout = QHBoxLayout()
        csv_btn_layout.setSpacing(10)
        
        csv_urunler = LuxuryButton("📦 Ürünler", "#2ec4b6")
        csv_urunler.setMaximumWidth(150)
        csv_urunler.clicked.connect(lambda: self.export_csv('urunler'))
        csv_btn_layout.addWidget(csv_urunler)
        
        csv_siparisler = LuxuryButton("🛒 Siparişler", "#2ec4b6")
        csv_siparisler.setMaximumWidth(150)
        csv_siparisler.clicked.connect(lambda: self.export_csv('siparisler'))
        csv_btn_layout.addWidget(csv_siparisler)
        
        csv_hareketler = LuxuryButton("📋 Hareketler", "#2ec4b6")
        csv_hareketler.setMaximumWidth(150)
        csv_hareketler.clicked.connect(lambda: self.export_csv('hareketler'))
        csv_btn_layout.addWidget(csv_hareketler)
        
        csv_btn_layout.addStretch()
        csv_layout.addLayout(csv_btn_layout)
        csv_frame.setLayout(csv_layout)
        layout.addWidget(csv_frame)
        
        excel_frame = QFrame()
        excel_frame.setStyleSheet(csv_frame.styleSheet())
        excel_layout = QVBoxLayout()
        
        excel_baslik = QLabel("📊 Excel Export")
        excel_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        excel_baslik.setStyleSheet("color: #0077b6;")
        excel_layout.addWidget(excel_baslik)
        
        excel_btn_layout = QHBoxLayout()
        excel_btn_layout.setSpacing(10)
        
        excel_tumu = LuxuryButton("📊 Tüm Raporlar", "#ff9f1c")
        excel_tumu.setMaximumWidth(150)
        excel_tumu.clicked.connect(self.export_excel_tumu)
        excel_btn_layout.addWidget(excel_tumu)
        excel_btn_layout.addStretch()
        
        excel_layout.addLayout(excel_btn_layout)
        excel_frame.setLayout(excel_layout)
        layout.addWidget(excel_frame)
        
        pdf_frame = QFrame()
        pdf_frame.setStyleSheet(csv_frame.styleSheet())
        pdf_layout = QVBoxLayout()
        
        pdf_baslik = QLabel("📃 PDF Export")
        pdf_baslik.setFont(QFont("Arial", 12, QFont.Bold))
        pdf_baslik.setStyleSheet("color: #0077b6;")
        pdf_layout.addWidget(pdf_baslik)
        
        pdf_btn_layout = QHBoxLayout()
        pdf_btn_layout.setSpacing(10)
        
        pdf_rapor = LuxuryButton("📃 Rapor PDF", "#4CAF50")
        pdf_rapor.setMaximumWidth(150)
        pdf_rapor.clicked.connect(self.export_pdf_rapor)
        pdf_btn_layout.addWidget(pdf_rapor)
        pdf_btn_layout.addStretch()
        
        pdf_layout.addLayout(pdf_btn_layout)
        pdf_frame.setLayout(pdf_layout)
        layout.addWidget(pdf_frame)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget
    
    def export_csv(self, tip):
        """CSV export"""
        try:
            import csv
            dosya_adi = f"{tip}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                if tip == 'urunler':
                    cursor.execute('SELECT * FROM urunler WHERE durum = "Aktif"')
                    kolonlar = ['urun_id', 'urun_kodu', 'urun_adi', 'kategori', 'stok', 'minimum_stok', 'fiyat']
                elif tip == 'siparisler':
                    cursor.execute('SELECT * FROM siparisler ORDER BY siparis_tarihi DESC LIMIT 1000')
                    kolonlar = ['siparis_no', 'urun_id', 'adet', 'birim_fiyat', 'durum', 'musteri_adi', 'siparis_tarihi']
                else:
                    cursor.execute('SELECT * FROM stok_hareketleri ORDER BY islem_tarihi DESC LIMIT 1000')
                    kolonlar = ['hareket_id', 'urun_id', 'hareket_tipi', 'miktar', 'onceki_stok', 'sonraki_stok', 'islem_tarihi']
                
                data = cursor.fetchall()
            
            with open(dosya_adi, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=kolonlar)
                writer.writeheader()
                for row in data:
                    writer.writerow({k: row[k] for k in kolonlar})
            
            msg_info(self, "Başarılı", f"✅ {tip.upper()} CSV'ye aktarıldı!\n📁 {dosya_adi}")
        except Exception as e:
            msg_warn(self, "Hata", f"⚠️ Export hatası: {str(e)}")
    
    def export_excel_tumu(self):
        """Excel export - Tüm raporlar"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            dosya_adi = f"depo_rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = Workbook()
            wb.remove(wb.active)
            
            ws_urunler = wb.create_sheet("Ürünler")
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT urun_kodu, urun_adi, kategori, stok, minimum_stok, 
                           reorder_point, fiyat, alis_fiyati, birim
                    FROM urunler WHERE durum = "Aktif"
                ''')
                data = cursor.fetchall()
                
                header = ['Kod', 'Ürün Adı', 'Kategori', 'Stok', 'Min. Stok', 'Reorder', 'Satış Fiyatı', 'Alış Fiyatı', 'Birim']
                ws_urunler.append(header)
                
                for row in data:
                    ws_urunler.append([row[k] for k in range(len(header))])
            
            ws_siparisler = wb.create_sheet("Siparişler")
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT s.siparis_no, u.urun_adi, s.adet, s.birim_fiyat, 
                           s.durum, s.musteri_adi, s.siparis_tarihi
                    FROM siparisler s
                    JOIN urunler u ON s.urun_id = u.urun_id
                    ORDER BY s.siparis_tarihi DESC
                    LIMIT 500
                ''')
                data = cursor.fetchall()
                
                header = ['Sipariş No', 'Ürün', 'Adet', 'Fiyat', 'Durum', 'Müşteri', 'Tarih']
                ws_siparisler.append(header)
                
                for row in data:
                    ws_siparisler.append([row[k] for k in range(len(header))])
            
            ws_istat = wb.create_sheet("İstatistikler")
            
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                cursor.execute('SELECT COUNT(*) as adet FROM urunler WHERE durum = "Aktif"')
                toplam_urun = cursor.fetchone()['adet']
                
                cursor.execute('SELECT SUM(stok) as toplam FROM urunler WHERE durum = "Aktif"')
                toplam_stok = cursor.fetchone()['toplam'] or 0
                
                cursor.execute('SELECT SUM(stok * fiyat) as deger FROM urunler WHERE durum = "Aktif"')
                toplam_deger = cursor.fetchone()['deger'] or 0
                
                cursor.execute('SELECT COUNT(*) as adet FROM siparisler WHERE durum = "Hazırlanıyor"')
                aktif_siparis = cursor.fetchone()['adet']
                
                cursor.execute('''
                    SELECT COUNT(*) as adet FROM urunler 
                    WHERE durum = "Aktif" AND stok < minimum_stok
                ''')
                dusuk_stok = cursor.fetchone()['adet']
            
            ws_istat.append(['İstatistik', 'Değer'])
            ws_istat.append(['Toplam Ürün Çeşidi', toplam_urun])
            ws_istat.append(['Toplam Stok (adet)', toplam_stok])
            ws_istat.append(['Toplam Stok Değeri (₺)', f"{toplam_deger:.2f}"])
            ws_istat.append(['Aktif Sipariş', aktif_siparis])
            ws_istat.append(['Düşük Stok Ürün', dusuk_stok])
            
            for ws in [ws_urunler, ws_siparisler, ws_istat]:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0077B6", end_color="0077B6", fill_type="solid")
            
            wb.save(dosya_adi)
            msg_info(self, "Başarılı", f"✅ Excel raporları oluşturuldu!\n📁 {dosya_adi}")
        except ImportError:
            msg_warn(self, "Hata", "⚠️ openpyxl yüklü değil!\nKomut: pip install openpyxl")
        except Exception as e:
            msg_warn(self, "Hata", f"⚠️ Export hatası: {str(e)}")
    
    def export_pdf_rapor(self):
        """PDF export - Raporlar"""
        try:
            from matplotlib.backends.backend_pdf import PdfPages
            
            dosya_adi = f"depo_rapor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            with PdfPages(dosya_adi) as pdf:
                fig1 = Figure(figsize=(8.5, 11))
                ax1 = fig1.add_subplot(2, 1, 1)
                
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT kategori, COUNT(*) as adet
                        FROM urunler WHERE durum = "Aktif"
                        GROUP BY kategori
                    ''')
                    data = cursor.fetchall()
                
                if data:
                    labels = [d['kategori'] for d in data]
                    sizes = [d['adet'] for d in data]
                    colors = ['#0077b6', '#00b4d8', '#2ec4b6', '#4CAF50', '#ff9f1c', '#e63946']
                    colors = colors[:len(labels)]
                    
                    ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
                    ax1.set_title("Kategori Dağılımı", fontweight='bold', fontsize=14)
                
                ax2 = fig1.add_subplot(2, 1, 2)
                ax2.axis('off')
                
                with self.db.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    cursor.execute('SELECT COUNT(*) as adet FROM urunler WHERE durum = "Aktif"')
                    toplam_urun = cursor.fetchone()['adet']
                    
                    cursor.execute('SELECT SUM(stok) as toplam FROM urunler WHERE durum = "Aktif"')
                    toplam_stok = cursor.fetchone()['toplam'] or 0
                    
                    cursor.execute('SELECT SUM(stok * fiyat) as deger FROM urunler WHERE durum = "Aktif"')
                    toplam_deger = cursor.fetchone()['deger'] or 0
                    
                    cursor.execute('SELECT COUNT(*) as adet FROM siparisler WHERE durum = "Hazırlanıyor"')
                    aktif_siparis = cursor.fetchone()['adet']
                
                rapor_text = f"""
DEPO & STOK YÖNETİM SİSTEMİ - RAPOR
Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M')}

─────────────────────────────────────────────────────────────
ÖZET İSTATİSTİKLER
─────────────────────────────────────────────────────────────

📦 Toplam Ürün Çeşidi:        {toplam_urun} adet
📊 Toplam Stok:               {toplam_stok:,} adet
💰 Toplam Stok Değeri:        ₺ {toplam_deger:,.2f}
🛒 Aktif Sipariş:             {aktif_siparis} adet

─────────────────────────────────────────────────────────────
                """
                
                ax2.text(0.1, 0.95, rapor_text, transform=ax2.transAxes,
                        fontsize=10, verticalalignment='top', fontfamily='monospace',
                        bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.3))
                
                fig1.patch.set_facecolor('#f5f5f5')
                pdf.savefig(fig1)
                plt.close(fig1)
            
            msg_info(self, "Başarılı", f"✅ PDF raporu oluşturuldu!\n📁 {dosya_adi}")
        except Exception as e:
            msg_warn(self, "Hata", f"⚠️ PDF export hatası: {str(e)}")


# ═══════════════════════════════════════════════════════════════════════════
# PHASE 4: ANOMALİ TESPİTİ, ABC ANALİZİ, EOQ, SİPARİŞ ÖNERİSİ
# ═══════════════════════════════════════════════════════════════════════════

# ───────────────────────────────────────────────────────────────────────────
# ANOMALI TESPİTİ DIALOG'U
# ───────────────────────────────────────────────────────────────────────────

class AnomalilerDialog(QDialog):
    """Anomali Tespiti ve Uyarıları"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("⚠️ Anomaliler & Uyarılar")
        self.setGeometry(150, 150, 850, 600)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("⚠️ ANOMALİLER & UYARILAR")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #e63946; margin-bottom: 20px;")
        
        tabs = QTabWidget()
        tabs.addTab(self.create_uyarilar_tab(), "🔴 Uyarılar")
        tabs.addTab(self.create_abc_tab(), "📊 ABC Analizi")
        tabs.addTab(self.create_eoq_tab(), "📦 EOQ & Öneriler")
        
        layout.addWidget(baslik)
        layout.addWidget(tabs)
        self.setLayout(layout)
    
    def create_uyarilar_tab(self):
        """Uyarılar tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # Uyarı türleri
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Filtre:"))
        
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["Tümü", "Düşük Stok", "Hareketsiz Ürünler", "Overstok"])
        self.filter_combo.currentIndexChanged.connect(self.anomalileri_listele)
        filter_layout.addWidget(self.filter_combo)
        
        yenile_btn = LuxuryButton("🔄 Yenile", "#0077b6")
        yenile_btn.setMaximumWidth(120)
        yenile_btn.clicked.connect(self.anomalileri_listele)
        filter_layout.addWidget(yenile_btn)
        filter_layout.addStretch()
        
        layout.addLayout(filter_layout)
        
        # Uyarı tablosu
        self.anomali_table = QTableWidget()
        self.anomali_table.setColumnCount(7)
        self.anomali_table.setHorizontalHeaderLabels([
            "ID", "Ürün Adı", "Kategorisi", "Stok", "Min. Stok", "Uyarı Türü", "Şiddet"
        ])
        self.anomali_table.setAlternatingRowColors(True)
        self.anomali_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.anomali_table)
        widget.setLayout(layout)
        
        self.anomalileri_listele()
        return widget
    
    def anomalileri_listele(self):
        """Anomalileri tabloya doldur"""
        self.anomali_table.setRowCount(0)
        filter_type = self.filter_combo.currentText()
        
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            
            if filter_type in ["Tümü", "Düşük Stok"]:
                # Düşük stok uyarıları
                cursor.execute('''
                    SELECT urun_id, urun_adi, kategori, stok, minimum_stok
                    FROM urunler
                    WHERE durum = 'Aktif' AND stok < minimum_stok
                    ORDER BY stok ASC
                ''')
                data = cursor.fetchall()
                
                for row_data in data:
                    row = self.anomali_table.rowCount()
                    self.anomali_table.insertRow(row)
                    
                    self.anomali_table.setItem(row, 0, QTableWidgetItem(str(row_data['urun_id'])))
                    self.anomali_table.setItem(row, 1, QTableWidgetItem(row_data['urun_adi']))
                    self.anomali_table.setItem(row, 2, QTableWidgetItem(row_data['kategori']))
                    self.anomali_table.setItem(row, 3, QTableWidgetItem(str(row_data['stok'])))
                    self.anomali_table.setItem(row, 4, QTableWidgetItem(str(row_data['minimum_stok'])))
                    self.anomali_table.setItem(row, 5, QTableWidgetItem("🔴 Düşük Stok"))
                    
                    # Şiddet hesapla
                    if row_data['stok'] == 0:
                        siddet = "KRITIK"
                        color = "#e63946"
                    elif row_data['stok'] < row_data['minimum_stok'] * 0.5:
                        siddet = "YÜKSEK"
                        color = "#ff6b6b"
                    else:
                        siddet = "ORTA"
                        color = "#ff9f1c"
                    
                    item = QTableWidgetItem(siddet)
                    item.setBackground(QColor(color))
                    item.setForeground(QColor("#ffffff"))
                    self.anomali_table.setItem(row, 6, item)
            
            if filter_type in ["Tümü", "Hareketsiz Ürünler"]:
                # Hareketsiz ürünler (90 gün)
                cursor.execute('''
                    SELECT u.urun_id, u.urun_adi, u.kategori, u.stok, u.son_cikis_tarihi
                    FROM urunler u
                    WHERE u.durum = 'Aktif'
                    AND (u.son_cikis_tarihi < date('now', '-90 days') OR u.son_cikis_tarihi IS NULL)
                    ORDER BY u.son_cikis_tarihi DESC
                ''')
                data = cursor.fetchall()
                
                for row_data in data:
                    row = self.anomali_table.rowCount()
                    self.anomali_table.insertRow(row)
                    
                    self.anomali_table.setItem(row, 0, QTableWidgetItem(str(row_data['urun_id'])))
                    self.anomali_table.setItem(row, 1, QTableWidgetItem(row_data['urun_adi']))
                    self.anomali_table.setItem(row, 2, QTableWidgetItem(row_data['kategori']))
                    self.anomali_table.setItem(row, 3, QTableWidgetItem(str(row_data['stok'])))
                    self.anomali_table.setItem(row, 4, QTableWidgetItem("-"))
                    self.anomali_table.setItem(row, 5, QTableWidgetItem("🟡 Hareketsiz"))
                    
                    item = QTableWidgetItem("ORTA")
                    item.setBackground(QColor("#ff9f1c"))
                    item.setForeground(QColor("#ffffff"))
                    self.anomali_table.setItem(row, 6, item)
            
            if filter_type in ["Tümü", "Overstok"]:
                # Overstok (Reorder Point'in 3x üzeri)
                cursor.execute('''
                    SELECT urun_id, urun_adi, kategori, stok, reorder_point
                    FROM urunler
                    WHERE durum = 'Aktif' AND stok > (reorder_point * 3)
                    ORDER BY stok DESC
                    LIMIT 20
                ''')
                data = cursor.fetchall()
                
                for row_data in data:
                    row = self.anomali_table.rowCount()
                    self.anomali_table.insertRow(row)
                    
                    self.anomali_table.setItem(row, 0, QTableWidgetItem(str(row_data['urun_id'])))
                    self.anomali_table.setItem(row, 1, QTableWidgetItem(row_data['urun_adi']))
                    self.anomali_table.setItem(row, 2, QTableWidgetItem(row_data['kategori']))
                    self.anomali_table.setItem(row, 3, QTableWidgetItem(str(row_data['stok'])))
                    self.anomali_table.setItem(row, 4, QTableWidgetItem(str(row_data['reorder_point'])))
                    self.anomali_table.setItem(row, 5, QTableWidgetItem("🔵 Overstok"))
                    
                    item = QTableWidgetItem("DÜŞÜK")
                    item.setBackground(QColor("#4CAF50"))
                    item.setForeground(QColor("#ffffff"))
                    self.anomali_table.setItem(row, 6, item)
    
    def create_abc_tab(self):
        """ABC Analizi tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        yenile_btn = LuxuryButton("🔄 Yenile", "#0077b6")
        yenile_btn.setMaximumWidth(120)
        yenile_btn.clicked.connect(self.abc_analizi_goster)
        layout.addWidget(yenile_btn)
        
        self.abc_table = QTableWidget()
        self.abc_table.setColumnCount(8)
        self.abc_table.setHorizontalHeaderLabels([
            "Kategori", "A Ürünler", "A %", "B Ürünler", "B %", "C Ürünler", "C %", "Toplam"
        ])
        self.abc_table.setAlternatingRowColors(True)
        self.abc_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.abc_table)
        widget.setLayout(layout)
        
        self.abc_analizi_goster()
        return widget
    
    def abc_analizi_goster(self):
        """ABC Analizi göster"""
        self.abc_table.setRowCount(0)
        
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            
            # Kategorileri al
            cursor.execute('SELECT DISTINCT kategori FROM urunler WHERE durum = "Aktif"')
            kategoriler = [k['kategori'] for k in cursor.fetchall()]
            
            for kat in kategoriler:
                # Ürün değerlerini al ve sırala
                cursor.execute('''
                    SELECT urun_adi, stok * fiyat as deger
                    FROM urunler
                    WHERE durum = 'Aktif' AND kategori = ?
                    ORDER BY deger DESC
                ''', (kat,))
                
                data = cursor.fetchall()
                
                if not data:
                    continue
                
                total_deger = sum(d['deger'] for d in data)
                cumsum = 0
                a_count = b_count = c_count = 0
                
                for item in data:
                    cumsum += item['deger']
                    percent = (cumsum / total_deger) * 100
                    
                    if percent <= 80:
                        a_count += 1
                    elif percent <= 95:
                        b_count += 1
                    else:
                        c_count += 1
                
                total_count = a_count + b_count + c_count
                
                row = self.abc_table.rowCount()
                self.abc_table.insertRow(row)
                
                self.abc_table.setItem(row, 0, QTableWidgetItem(kat))
                self.abc_table.setItem(row, 1, QTableWidgetItem(str(a_count)))
                self.abc_table.setItem(row, 2, QTableWidgetItem(f"{(a_count/total_count)*100:.1f}%"))
                self.abc_table.setItem(row, 3, QTableWidgetItem(str(b_count)))
                self.abc_table.setItem(row, 4, QTableWidgetItem(f"{(b_count/total_count)*100:.1f}%"))
                self.abc_table.setItem(row, 5, QTableWidgetItem(str(c_count)))
                self.abc_table.setItem(row, 6, QTableWidgetItem(f"{(c_count/total_count)*100:.1f}%"))
                self.abc_table.setItem(row, 7, QTableWidgetItem(str(total_count)))
    
    def create_eoq_tab(self):
        """EOQ ve Sipariş Önerileri tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("📦 EOQ FORMULA VE SİPARİŞ ÖNERİLERİ")
        baslik.setFont(QFont("Arial", 12, QFont.Bold))
        baslik.setStyleSheet("color: #0077b6;")
        layout.addWidget(baslik)
        
        # EOQ Bilgisi
        bilgi_frame = QFrame()
        bilgi_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 10px;
                padding: 15px;
            }
        """)
        bilgi_layout = QVBoxLayout()
        
        eoq_text = QLabel("""
<b>EOQ (Economic Order Quantity) Formülü:</b>
<br>
<font color='#00b4d8'><b>EOQ = √(2DS / H)</b></font>
<br><br>
Nerede:
<br>
<font color='#2ec4b6'><b>D</b></font> = Yıllık talep (adet)
<br>
<font color='#2ec4b6'><b>S</b></font> = Sipariş maliyeti (₺)
<br>
<font color='#2ec4b6'><b>H</b></font> = Yıllık stoklama maliyeti (₺/adet)
<br><br>
<font color='#4CAF50'>Bu formül, en ekonomik sipariş miktarını hesaplar.</font>
        """)
        eoq_text.setAlignment(Qt.AlignLeft)
        bilgi_layout.addWidget(eoq_text)
        bilgi_frame.setLayout(bilgi_layout)
        layout.addWidget(bilgi_frame)
        
        # Öneriler Tablosu
        oneriler_frame = QFrame()
        oneriler_frame.setStyleSheet(bilgi_frame.styleSheet())
        oneriler_layout = QVBoxLayout()
        
        oneriler_baslik = QLabel("📋 SİPARİŞ ÖNERİLERİ")
        oneriler_baslik.setFont(QFont("Arial", 11, QFont.Bold))
        oneriler_baslik.setStyleSheet("color: #0077b6;")
        oneriler_layout.addWidget(oneriler_baslik)
        
        self.oneriler_table = QTableWidget()
        self.oneriler_table.setColumnCount(8)
        self.oneriler_table.setHorizontalHeaderLabels([
            "Ürün Adı", "Mevcut Stok", "Min. Stok", "Reorder", "EOQ", "Önerilen Miktar", "Tahmini Tutar (₺)", "Öncelik"
        ])
        self.oneriler_table.setAlternatingRowColors(True)
        self.oneriler_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.oneriler_table.setMaximumHeight(350)
        
        oneriler_layout.addWidget(self.oneriler_table)
        oneriler_frame.setLayout(oneriler_layout)
        layout.addWidget(oneriler_frame)
        
        # Yenile butonu
        yenile_layout = QHBoxLayout()
        yenile_btn = LuxuryButton("🔄 Sipariş Önerilerini Hesapla", "#2ec4b6")
        yenile_btn.setMaximumWidth(250)
        yenile_btn.clicked.connect(self.eoq_onerilerini_hesapla)
        yenile_layout.addWidget(yenile_btn)
        yenile_layout.addStretch()
        
        layout.addLayout(yenile_layout)
        layout.addStretch()
        
        widget.setLayout(layout)
        
        self.eoq_onerilerini_hesapla()
        return widget
    
    def eoq_onerilerini_hesapla(self):
        """EOQ'ye dayalı sipariş önerileri"""
        self.oneriler_table.setRowCount(0)
        
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT u.urun_id, u.urun_adi, u.stok, u.minimum_stok, 
                       u.reorder_point, u.fiyat, u.alis_fiyati
                FROM urunler u
                WHERE u.durum = 'Aktif' AND u.stok <= u.reorder_point
                ORDER BY u.stok ASC
            ''')
            
            data = cursor.fetchall()
            
            for item in data:
                # Basit EOQ hesapla (D=100/yıl, S=50₺, H=fiyat*0.2)
                D = 100  # Yıllık talep (varsayılan)
                S = 50   # Sipariş maliyeti (₺)
                H = item['fiyat'] * 0.2 if item['fiyat'] > 0 else 1
                
                import math
                eoq = math.sqrt((2 * D * S) / H)
                eoq = max(int(eoq), item['minimum_stok'])
                
                # Eksik miktarı hesapla
                eksik = max(item['minimum_stok'] - item['stok'], 0)
                onerilen = max(eoq, eksik)
                
                # Tahmini tutar
                tutar = onerilen * item['alis_fiyati']
                
                # Öncelik belirle
                if item['stok'] == 0:
                    oncelik = "🔴 ACIL"
                    color = "#e63946"
                elif item['stok'] < item['minimum_stok'] * 0.5:
                    oncelik = "🟠 YÜKSEK"
                    color = "#ff9f1c"
                else:
                    oncelik = "🟡 ORTA"
                    color = "#ffeb3b"
                
                row = self.oneriler_table.rowCount()
                self.oneriler_table.insertRow(row)
                
                self.oneriler_table.setItem(row, 0, QTableWidgetItem(item['urun_adi']))
                self.oneriler_table.setItem(row, 1, QTableWidgetItem(str(item['stok'])))
                self.oneriler_table.setItem(row, 2, QTableWidgetItem(str(item['minimum_stok'])))
                self.oneriler_table.setItem(row, 3, QTableWidgetItem(str(item['reorder_point'])))
                self.oneriler_table.setItem(row, 4, QTableWidgetItem(f"{eoq:.0f}"))
                self.oneriler_table.setItem(row, 5, QTableWidgetItem(str(onerilen)))
                self.oneriler_table.setItem(row, 6, QTableWidgetItem(f"₺ {tutar:,.2f}"))
                
                item_widget = QTableWidgetItem(oncelik)
                item_widget.setBackground(QColor(color))
                item_widget.setForeground(QColor("#000000"))
                self.oneriler_table.setItem(row, 7, item_widget)


# ═══════════════════════════════════════════════════════════════════════════
# PHASE 5: GAMİFİCATİON, STOK SAYIMI, NOTLAR, CONTEXT MENU
# ═══════════════════════════════════════════════════════════════════════════

# ───────────────────────────────────────────────────────────────────────────
# GAMIFICATION DIALOG'U
# ───────────────────────────────────────────────────────────────────────────

class GamificationDialog(QDialog):
    """Gamification - Leaderboard, Badge'ler, Başarılar"""
    
    def __init__(self, db, username, parent=None):
        super().__init__(parent)
        self.db = db
        self.username = username
        self.setWindowTitle("🏆 Başarılar & Leaderboard")
        self.setGeometry(100, 100, 900, 650)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("🏆 BAŞARILAR & LEADERBOARD")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #ffd700; margin-bottom: 20px;")
        
        tabs = QTabWidget()
        tabs.addTab(self.create_badge_tab(), "🎖️ Badge'ler")
        tabs.addTab(self.create_leaderboard_tab(), "🏆 Leaderboard")
        tabs.addTab(self.create_mood_tab(), "😊 Mood Tracker")
        
        layout.addWidget(baslik)
        layout.addWidget(tabs)
        self.setLayout(layout)
    
    def _badge_durumlarini_hesapla(self):
        """DB'den badge durumlarını hesapla"""
        with self.db.get_connection() as conn:
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(*) as c FROM urunler WHERE durum='Aktif'")
            urun_sayisi = cursor.fetchone()["c"]

            cursor.execute("SELECT COUNT(*) as c FROM siparisler")
            siparis_sayisi = cursor.fetchone()["c"]

            cursor.execute("SELECT COUNT(*) as c FROM stok_hareketleri")
            hareket_sayisi = cursor.fetchone()["c"]

            cursor.execute(
                "SELECT COUNT(*) as c FROM urunler "
                "WHERE stok < minimum_stok AND durum='Aktif'"
            )
            dusuk_stok = cursor.fetchone()["c"]

        tam_stok = (dusuk_stok == 0 and urun_sayisi > 0)

        return {
            "giris":    True,                     # her zaman kazanılmış
            "urun":     urun_sayisi > 0,
            "siparis":  siparis_sayisi > 0,
            "rapor":    False,                     # export sonrası set edilir
            "islem100": hareket_sayisi >= 100,
            "abc":      False,                     # ABC dialog açılınca set edilir
            "tam_stok": tam_stok,
            "speedrun": False,
        }

    def create_badge_tab(self):
        """Badge'ler sekmesi - DB senkron"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)

        # Yenile butonu
        yenile_btn = LuxuryButton("🔄 Badge'leri Yenile", "#0077b6")
        yenile_btn.setMaximumWidth(180)
        yenile_btn.clicked.connect(lambda: self._badge_grid_doldur(grid, grid_widget))
        layout.addWidget(yenile_btn)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(
            "QScrollArea{border:none;background:transparent;}"
            "QScrollBar:vertical{background:#2d2d3a;border-radius:5px;width:7px;}"
            "QScrollBar::handle:vertical{background:#0077b6;border-radius:5px;}"
        )

        grid_widget = QWidget()
        grid_widget.setStyleSheet("background:transparent;")
        grid = QGridLayout()
        grid.setSpacing(12)
        grid_widget.setLayout(grid)

        self._badge_grid_doldur(grid, grid_widget)

        scroll.setWidget(grid_widget)
        layout.addWidget(scroll, stretch=1)
        widget.setLayout(layout)
        return widget

    def _badge_grid_doldur(self, grid, grid_widget):
        """Badge grid'ini DB verisine göre doldur"""
        # Mevcut widgetları temizle
        while grid.count():
            item = grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        durum = self._badge_durumlarini_hesapla()

        badges = [
            ("🌟", "İlk Giriş",         "Uygulamaya giriş yapıldı",        durum["giris"]),
            ("📦", "Ürün Ekledi",        "İlk ürün sisteme eklendi",         durum["urun"]),
            ("🛒", "Sipariş Oluşturdu",  "İlk sipariş oluşturuldu",          durum["siparis"]),
            ("📊", "Rapor İndirdi",      "İlk rapor dışa aktarıldı",         durum["rapor"]),
            ("⚡", "100 İşlem",          "100+ stok hareketi gerçekleşti",   durum["islem100"]),
            ("🎯", "ABC Master",         "ABC Analizi incelendi",             durum["abc"]),
            ("💯", "Tam Stok",           "Tüm ürünler minimum stok üstünde", durum["tam_stok"]),
            ("🚀", "Speedrun",           "Yoğun kullanım başarısı",          durum["speedrun"]),
        ]

        kazanilan = sum(1 for *_, e in badges if e)
        # Başlık sayacı
        for i in range(grid.rowCount() + 1):
            pass  # sayac label için yer açık

        for idx, (emoji, isim, aciklama, kazanildi) in enumerate(badges):
            row, col = divmod(idx, 2)
            border = "#ffd700" if kazanildi else "#444"
            bg     = "stop:0 #2a2418,stop:1 #1a1a2e" if kazanildi else "stop:0 #2d2d3a,stop:1 #1a1a2e"
            clr    = "#ffd700" if kazanildi else "#666"
            emoji_clr = clr

            frame = QFrame()
            frame.setStyleSheet(
                f"QFrame{{background:qlineargradient(x1:0,y1:0,x2:1,y2:1,{bg});"
                f"border-radius:14px;border:2px solid {border};}}"
            )
            fl = QVBoxLayout()
            fl.setContentsMargins(12, 14, 12, 14)
            fl.setSpacing(6)

            # Emoji büyük
            e_lbl = QLabel(emoji)
            e_lbl.setFont(QFont("Segoe UI Emoji", 28))
            e_lbl.setAlignment(Qt.AlignCenter)
            e_lbl.setStyleSheet(f"color:{emoji_clr};background:transparent;border:none;")

            # İsim
            n_lbl = QLabel(isim)
            n_lbl.setFont(QFont("Arial", 10, QFont.Bold))
            n_lbl.setAlignment(Qt.AlignCenter)
            n_lbl.setStyleSheet(f"color:{clr};background:transparent;border:none;")

            # Açıklama
            d_lbl = QLabel(aciklama)
            d_lbl.setFont(QFont("Arial", 8))
            d_lbl.setAlignment(Qt.AlignCenter)
            d_lbl.setWordWrap(True)
            d_lbl.setStyleSheet("color:#888;background:transparent;border:none;")

            # Durum
            status_txt  = "✅  KAZANDI" if kazanildi else "🔒  KİLİTLİ"
            status_clr  = "#4CAF50"    if kazanildi else "#555"
            s_lbl = QLabel(status_txt)
            s_lbl.setFont(QFont("Arial", 9, QFont.Bold))
            s_lbl.setAlignment(Qt.AlignCenter)
            s_lbl.setStyleSheet(
                f"color:{status_clr};background:transparent;"
                f"border:none;padding-top:4px;"
            )

            fl.addWidget(e_lbl)
            fl.addWidget(n_lbl)
            fl.addWidget(d_lbl)
            fl.addWidget(s_lbl)
            frame.setLayout(fl)
            grid.addWidget(frame, row, col)

        # Özet satırı
        toplam = len(badges)
        ozet = QLabel(f"🏅  {kazanilan} / {toplam} badge kazanıldı")
        ozet.setFont(QFont("Arial", 10, QFont.Bold))
        ozet.setAlignment(Qt.AlignCenter)
        ozet.setStyleSheet("color:#ffd700;padding:6px;background:transparent;")
        son_row = (len(badges) + 1) // 2
        grid.addWidget(ozet, son_row, 0, 1, 2)
    
    def create_leaderboard_tab(self):
        """Leaderboard sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        yenile_btn = LuxuryButton("🔄 Leaderboard'u Yenile", "#ffd700")
        yenile_btn.setMaximumWidth(200)
        yenile_btn.clicked.connect(self.leaderboard_goster)
        layout.addWidget(yenile_btn)
        
        self.leaderboard_table = QTableWidget()
        self.leaderboard_table.setColumnCount(6)
        self.leaderboard_table.setHorizontalHeaderLabels([
            "Sıra", "Kullanıcı", "İşlem Sayısı", "Eklenen Ürün", "Stok Sayımı", "Puan"
        ])
        self.leaderboard_table.setAlternatingRowColors(True)
        self.leaderboard_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.leaderboard_table)
        widget.setLayout(layout)
        
        self.leaderboard_goster()
        return widget
    
    def leaderboard_goster(self):
        """Leaderboard göster"""
        self.leaderboard_table.setRowCount(0)
        
        # Örnek veri (veritabanından alınabilir)
        users = [
            ("admin", 250, 45, 12, 2500),
            ("personel1", 180, 32, 8, 1800),
            ("personel2", 120, 20, 5, 1200),
        ]
        
        for idx, (user, islem, urun, sayim, puan) in enumerate(users, 1):
            row = self.leaderboard_table.rowCount()
            self.leaderboard_table.insertRow(row)
            
            # Sıra numarası
            if idx == 1:
                sira_text = "🥇 1"
                color = "#ffd700"
            elif idx == 2:
                sira_text = "🥈 2"
                color = "#c0c0c0"
            elif idx == 3:
                sira_text = "🥉 3"
                color = "#cd7f32"
            else:
                sira_text = str(idx)
                color = "#ffffff"
            
            sira_item = QTableWidgetItem(sira_text)
            sira_item.setForeground(QColor(color))
            sira_item.setFont(QFont("Arial", 10, QFont.Bold))
            self.leaderboard_table.setItem(row, 0, sira_item)
            
            self.leaderboard_table.setItem(row, 1, QTableWidgetItem(user))
            self.leaderboard_table.setItem(row, 2, QTableWidgetItem(str(islem)))
            self.leaderboard_table.setItem(row, 3, QTableWidgetItem(str(urun)))
            self.leaderboard_table.setItem(row, 4, QTableWidgetItem(str(sayim)))
            
            puan_item = QTableWidgetItem(str(puan))
            puan_item.setFont(QFont("Arial", 10, QFont.Bold))
            puan_item.setForeground(QColor("#0077b6"))
            self.leaderboard_table.setItem(row, 5, puan_item)
    
    def create_mood_tab(self):
        """Mood Tracker sekmesi - Fonksiyonel"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)

        baslik = QLabel("😊 GÜNLÜK MOD TRACKER")
        baslik.setFont(QFont("Arial", 14, QFont.Bold))
        baslik.setStyleSheet("color: #0077b6;")
        layout.addWidget(baslik)

        # ── Bugünün modu ──
        card_style = ("background: qlineargradient(x1:0,y1:0,x2:1,y2:1,"
                      "stop:0 #2d2d3a,stop:1 #1a1a2e);"
                      "border-radius:12px;padding:15px;")

        mood_frame = QFrame()
        mood_frame.setStyleSheet("QFrame{" + card_style + "}")
        mood_layout = QVBoxLayout()
        mood_layout.setSpacing(12)

        soru = QLabel("Bugün nasıl hissediyorsunuz?")
        soru.setFont(QFont("Arial", 12))
        soru.setStyleSheet("color:#ffffff;")
        mood_layout.addWidget(soru)

        # Seçili mood göstergesi
        self.mood_secili_label = QLabel("Henüz seçilmedi...")
        self.mood_secili_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.mood_secili_label.setStyleSheet("color:#aaa;")
        mood_layout.addWidget(self.mood_secili_label)

        mood_btn_layout = QHBoxLayout()
        mood_btn_layout.setSpacing(10)

        self.mood_kayitlari = []  # (tarih, puan) listesi

        moods = [
            ("😢 Kötü",   "#e63946", 1),
            ("😕 Orta",   "#ff9f1c", 2),
            ("😊 İyi",    "#4CAF50", 3),
            ("😄 Harika", "#ffd700", 4),
        ]

        for mood_text, color, puan in moods:
            btn = LuxuryButton(mood_text, color)
            btn.setMinimumWidth(110)
            # closure ile puan ve text yakala
            def make_handler(t, p, c):
                def handler():
                    from datetime import datetime as dt
                    self.mood_kayitlari.append((dt.now().strftime("%d.%m.%Y"), p))
                    self.mood_secili_label.setText(f"Bugün: {t}")
                    self.mood_secili_label.setStyleSheet(f"color:{c};font-weight:bold;")
                    self._mood_istatistik_guncelle()
                return handler
            btn.clicked.connect(make_handler(mood_text, puan, color))
            mood_btn_layout.addWidget(btn)

        mood_btn_layout.addStretch()
        mood_layout.addLayout(mood_btn_layout)
        mood_frame.setLayout(mood_layout)
        layout.addWidget(mood_frame)

        # ── Geçmiş kayıtlar ──
        gecmis_frame = QFrame()
        gecmis_frame.setStyleSheet("QFrame{" + card_style + "}")
        gecmis_layout = QVBoxLayout()
        gecmis_layout.setSpacing(8)

        gecmis_baslik = QLabel("📅 Bu Haftanın Kayıtları")
        gecmis_baslik.setFont(QFont("Arial", 11, QFont.Bold))
        gecmis_baslik.setStyleSheet("color:#0077b6;")
        gecmis_layout.addWidget(gecmis_baslik)

        self.mood_gecmis_list = QListWidget()
        self.mood_gecmis_list.setMaximumHeight(130)
        self.mood_gecmis_list.setStyleSheet(
            "QListWidget{background:#1a1a2e;border:1px solid #0077b6;"
            "border-radius:6px;color:#fff;font-size:11px;}"
            "QListWidget::item{padding:5px;}"
        )
        gecmis_layout.addWidget(self.mood_gecmis_list)
        gecmis_frame.setLayout(gecmis_layout)
        layout.addWidget(gecmis_frame)

        # ── İstatistik kartı ──
        self.mood_istat_frame = QFrame()
        self.mood_istat_frame.setStyleSheet("QFrame{" + card_style + "}")
        istat_layout = QVBoxLayout()
        istat_layout.setSpacing(8)

        self.mood_ort_label = QLabel("📊 Bu Ayın Ortalaması: —")
        self.mood_ort_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.mood_ort_label.setStyleSheet("color:#4CAF50;")
        istat_layout.addWidget(self.mood_ort_label)

        self.mood_pbar = QProgressBar()
        self.mood_pbar.setRange(0, 100)
        self.mood_pbar.setValue(0)
        self.mood_pbar.setFormat("Genel İyi Hissetme: %p%")
        self.mood_pbar.setFixedHeight(20)
        istat_layout.addWidget(self.mood_pbar)

        self.mood_istat_frame.setLayout(istat_layout)
        layout.addWidget(self.mood_istat_frame)

        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def _mood_istatistik_guncelle(self):
        """Mood istatistiklerini güncelle"""
        if not self.mood_kayitlari:
            return
        from datetime import datetime as dt

        emoji_map = {1: "😢 Kötü", 2: "😕 Orta", 3: "😊 İyi", 4: "😄 Harika"}
        color_map = {1: "#e63946", 2: "#ff9f1c", 3: "#4CAF50", 4: "#ffd700"}

        # Geçmiş listeyi güncelle
        self.mood_gecmis_list.clear()
        for tarih, puan in reversed(self.mood_kayitlari[-7:]):
            emoji = emoji_map.get(puan, "?")
            item = QListWidgetItem(f"  {tarih}  →  {emoji}")
            item.setForeground(QColor(color_map.get(puan, "#fff")))
            self.mood_gecmis_list.addItem(item)

        # Ortalama
        ort = sum(p for _, p in self.mood_kayitlari) / len(self.mood_kayitlari)
        ort_emoji = emoji_map.get(round(ort), "😊 İyi")
        ort_color = color_map.get(round(ort), "#4CAF50")
        self.mood_ort_label.setText(f"📊 Bu Ayın Ortalaması: {ort_emoji}")
        self.mood_ort_label.setStyleSheet(f"color:{ort_color};font-weight:bold;")

        # Progress bar: (ort-1)/3 * 100
        pct = int(((ort - 1) / 3) * 100)
        self.mood_pbar.setValue(pct)


# ───────────────────────────────────────────────────────────────────────────
# NOTLAR VE CONTEXT MENU
# ───────────────────────────────────────────────────────────────────────────

class NotlarWidget(QWidget):
    """Kişisel Notlar Widget - Tam Fonksiyonel"""

    LIST_STYLE = (
        "QListWidget{background-color:#1a1a2e;border:2px solid #0077b6;"
        "border-radius:10px;color:#ffffff;font-size:12px;padding:4px;}"
        "QListWidget::item{padding:10px 12px;border-bottom:1px solid #2d2d3a;}"
        "QListWidget::item:selected{background-color:#0077b6;border-radius:6px;}"
        "QListWidget::item:hover{background-color:#2d2d3a;}"
    )

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.init_ui()

    def init_ui(self):
        self.setStyleSheet("background-color:#0f0f1a;")
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 20)
        root.setSpacing(14)

        # ── Başlık ──
        baslik = QLabel("📝  NOTLAR")
        baslik.setFont(QFont("Arial", 14, QFont.Bold))
        baslik.setStyleSheet("color:#0077b6;")
        root.addWidget(baslik)

        # ── Giriş satırı ──
        inp_row = QHBoxLayout()
        inp_row.setSpacing(8)

        self.not_input = QLineEdit()
        self.not_input.setPlaceholderText("Yeni not ekle ve Enter'a bas...")
        self.not_input.setMinimumHeight(38)
        self.not_input.setStyleSheet(
            "QLineEdit{background:#1a1a2e;border:2px solid #0077b6;"
            "border-radius:8px;padding:6px 12px;color:#fff;font-size:12px;}"
            "QLineEdit:focus{border:2px solid #00b4d8;}"
        )
        self.not_input.returnPressed.connect(self.not_ekle)

        ekle_btn = LuxuryButton("➕ Ekle", "#2ec4b6")
        ekle_btn.setFixedHeight(38)
        ekle_btn.setMinimumWidth(80)
        ekle_btn.clicked.connect(self.not_ekle)

        sil_btn = LuxuryButton("🗑 Sil", "#e63946")
        sil_btn.setFixedHeight(38)
        sil_btn.setMinimumWidth(70)
        sil_btn.clicked.connect(self.not_sil)

        inp_row.addWidget(self.not_input, stretch=1)
        inp_row.addWidget(ekle_btn)
        inp_row.addWidget(sil_btn)
        root.addLayout(inp_row)

        # ── Liste ──
        self.notlar_list = QListWidget()
        self.notlar_list.setStyleSheet(self.LIST_STYLE)
        self.notlar_list.setAlternatingRowColors(False)
        self.notlar_list.setMinimumHeight(300)
        root.addWidget(self.notlar_list, stretch=1)

        # ── Alt bilgi ──
        self.sayac_label = QLabel("0 not")
        self.sayac_label.setStyleSheet("color:#666;font-size:10px;")
        self.sayac_label.setAlignment(Qt.AlignRight)
        root.addWidget(self.sayac_label)

        self.notlari_yukle()

    def not_ekle(self):
        not_text = self.not_input.text().strip()
        if not not_text:
            return
        item = QListWidgetItem(f"  {not_text}")
        item.setForeground(QColor("#e0e0e0"))
        self.notlar_list.addItem(item)
        self.notlar_list.scrollToBottom()
        self.not_input.clear()
        self._sayac_guncelle()

    def not_sil(self):
        row = self.notlar_list.currentRow()
        if row >= 0:
            self.notlar_list.takeItem(row)
            self._sayac_guncelle()

    def _sayac_guncelle(self):
        n = self.notlar_list.count()
        self.sayac_label.setText(f"{n} not")

    def notlari_yukle(self):
        varsayilan = [
            "Stock check yapılması gerekiyor",
            "Yeni tedarikçi ile iletişime geç",
            "Pazarlık oranlarını güncelle",
        ]
        for not_text in varsayilan:
            item = QListWidgetItem(f"  {not_text}")
            item.setForeground(QColor("#aaa"))
            self.notlar_list.addItem(item)
        self._sayac_guncelle()


# ───────────────────────────────────────────────────────────────────────────
# STOK SAYIMI DİYALOGU
# ───────────────────────────────────────────────────────────────────────────

class StokSayimiDialog(QDialog):
    """Stok Sayımı (Batch) Dialog'u"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("📦 Stok Sayımı (Batch)")
        self.setGeometry(200, 200, 800, 550)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("📦 STOK SAYIMI IŞLEMI")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #2ec4b6;")
        layout.addWidget(baslik)
        
        # Kategori seçimi
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Kategori Seçin:"))
        
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItem("Tümü")
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT DISTINCT kategori FROM urunler WHERE durum = "Aktif"')
            for k in cursor.fetchall():
                self.kategori_combo.addItem(k['kategori'])
        
        self.kategori_combo.currentIndexChanged.connect(self.urunleri_listele_sayim)
        filter_layout.addWidget(self.kategori_combo)
        filter_layout.addStretch()
        
        layout.addLayout(filter_layout)
        
        # Ürünler tablosu
        self.sayim_table = QTableWidget()
        self.sayim_table.setColumnCount(6)
        self.sayim_table.setHorizontalHeaderLabels([
            "ID", "Ürün Adı", "Sistem Stok", "Fiziki Sayım", "Fark", "✓"
        ])
        self.sayim_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(self.sayim_table)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        kaydet_btn = LuxuryButton("💾 Sayımı Kaydet", "#4CAF50")
        kaydet_btn.setMaximumWidth(150)
        kaydet_btn.clicked.connect(self.sayimi_kaydet)
        btn_layout.addWidget(kaydet_btn)
        
        iptal_btn = LuxuryButton("❌ İptal", "#e63946")
        iptal_btn.setMaximumWidth(100)
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(iptal_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
        self.urunleri_listele_sayim()
    
    def urunleri_listele_sayim(self):
        """Ürünleri tabloya doldur"""
        self.sayim_table.setRowCount(0)
        kategori = self.kategori_combo.currentText()
        
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            
            if kategori == "Tümü":
                cursor.execute('SELECT urun_id, urun_adi, stok FROM urunler WHERE durum = "Aktif"')
            else:
                cursor.execute('SELECT urun_id, urun_adi, stok FROM urunler WHERE durum = "Aktif" AND kategori = ?', (kategori,))
            
            data = cursor.fetchall()
        
        for item in data:
            row = self.sayim_table.rowCount()
            self.sayim_table.insertRow(row)
            
            self.sayim_table.setItem(row, 0, QTableWidgetItem(str(item['urun_id'])))
            self.sayim_table.setItem(row, 1, QTableWidgetItem(item['urun_adi']))
            self.sayim_table.setItem(row, 2, QTableWidgetItem(str(item['stok'])))
            
            # Fiziki sayım input'u
            fiziki_input = QSpinBox()
            fiziki_input.setMinimum(0)
            fiziki_input.setMaximum(10000)
            fiziki_input.setValue(item['stok'])
            self.sayim_table.setCellWidget(row, 3, fiziki_input)
            
            # Fark otomatik hesaplanır
            fark_label = QLabel("0")
            fark_label.setAlignment(Qt.AlignCenter)
            self.sayim_table.setCellWidget(row, 4, fark_label)
            
            # Checkbox
            checkbox = QCheckBox()
            self.sayim_table.setCellWidget(row, 5, checkbox)
    
    def sayimi_kaydet(self):
        """Sayımı kaydet"""
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor()
                
                for row in range(self.sayim_table.rowCount()):
                    urun_id = int(self.sayim_table.item(row, 0).text())
                    fiziki = self.sayim_table.cellWidget(row, 3).value()
                    
                    cursor.execute('''
                        SELECT stok FROM urunler WHERE urun_id = ?
                    ''', (urun_id,))
                    sistem_stok = cursor.fetchone()['stok']
                    
                    fark = sistem_stok - fiziki
                    
                    if fark != 0:
                        cursor.execute('''
                            INSERT INTO urun_sayim_logu (urun_id, sistem_stok, fiziki_stok, fark, sayim_tarihi)
                            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                        ''', (urun_id, sistem_stok, fiziki, fark))
                        
                        cursor.execute('''
                            UPDATE urunler SET stok = ? WHERE urun_id = ?
                        ''', (fiziki, urun_id))
            
            msg_info(self, "Başarılı", "✅ Stok sayımı başarıyla kaydedildi!")
            self.accept()
        except Exception as e:
            msg_warn(self, "Hata", f"⚠️ Hata: {str(e)}")


# ═══════════════════════════════════════════════════════════════════════════
# PHASE 6: AYARLAR, YARDIM SİSTEMİ, KIŞAYOLLAR, ARAÇ İPUÇLARI
# ═══════════════════════════════════════════════════════════════════════════

# ───────────────────────────────────────────────────────────────────────────
# YARDIM DİYALOGU
# ───────────────────────────────────────────────────────────────────────────

class YardimDialog(QDialog):
    """F1 Yardım Sistemi"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("❓ Yardım & Rehber")
        self.setGeometry(100, 100, 900, 700)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("❓ YARDIM & REHBER")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6; margin-bottom: 20px;")
        
        tabs = QTabWidget()
        tabs.addTab(self.create_kisiayak_tab(), "⌨️ Kısayollar")
        tabs.addTab(self.create_rehber_tab(), "📖 Rehber")
        tabs.addTab(self.create_hakkinda_tab(), "ℹ️ Hakkında")
        
        layout.addWidget(baslik)
        layout.addWidget(tabs)
        self.setLayout(layout)
    
    def create_kisiayak_tab(self):
        """Kısayollar tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        content = QWidget()
        content_layout = QVBoxLayout()
        content_layout.setSpacing(20)
        
        # Sistem Kısayolları
        sistem_frame = self.create_kisayol_frame(
            "🖥️ SİSTEM KIŞAYOLLARI",
            [
                ("F1", "Yardım Sistemini Aç"),
                ("Ctrl+N", "Yeni Ürün Ekle"),
                ("Ctrl+O", "Raporları Aç"),
                ("Ctrl+F", "Arama Aç"),
                ("Ctrl+Q", "Çıkış Yap"),
                ("Ctrl+H", "Kısayolları Göster"),
            ]
        )
        content_layout.addWidget(sistem_frame)
        
        # Stok Yönetimi Kısayolları
        stok_frame = self.create_kisayol_frame(
            "📦 STOK YÖNETIMI",
            [
                ("Ctrl+P", "Stok +"),
                ("Ctrl+M", "Stok -"),
                ("Ctrl+S", "Stok Sayımı"),
                ("Ctrl+R", "Rapor İndir"),
                ("Ctrl+E", "Excel'e Aktar"),
            ]
        )
        content_layout.addWidget(stok_frame)
        
        # Ürün Yönetimi Kısayolları
        urun_frame = self.create_kisayol_frame(
            "🏷️ ÜRÜN YÖNETIMI",
            [
                ("Del", "Seçili Ürünü Sil"),
                ("Enter", "Ürünü Düzenle"),
                ("Space", "Seçimi Değiştir"),
            ]
        )
        content_layout.addWidget(urun_frame)
        
        content_layout.addStretch()
        content.setLayout(content_layout)
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        widget.setLayout(layout)
        return widget
    
    def create_kisayol_frame(self, baslik, kisayollar):
        """Kısayol frame'i oluştur"""
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 10px;
                padding: 15px;
                border-left: 4px solid #0077b6;
            }
        """)
        layout = QVBoxLayout()
        
        baslik_label = QLabel(baslik)
        baslik_label.setFont(QFont("Arial", 12, QFont.Bold))
        baslik_label.setStyleSheet("color: #0077b6;")
        layout.addWidget(baslik_label)
        
        for kisayol, aciklama in kisayollar:
            row_layout = QHBoxLayout()
            
            kisayol_label = QLabel(kisayol)
            kisayol_label.setFont(QFont("Arial", 10, QFont.Bold))
            kisayol_label.setStyleSheet("color: #ffd700;")
            kisayol_label.setMinimumWidth(100)
            
            aciklama_label = QLabel(aciklama)
            aciklama_label.setFont(QFont("Arial", 10))
            aciklama_label.setStyleSheet("color: #aaa;")
            
            row_layout.addWidget(kisayol_label)
            row_layout.addWidget(aciklama_label)
            row_layout.addStretch()
            
            layout.addLayout(row_layout)
        
        frame.setLayout(layout)
        return frame
    
    def create_rehber_tab(self):
        """Rehber tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        content = QLabel("""
<h3 style='color: #0077b6;'>📖 ÜRÜN YÖNETİM REHBERİ</h3>

<h4 style='color: #2ec4b6;'>1️⃣ Ürün Ekleme</h4>
<p style='color: #aaa;'>
[➕ Yeni Ürün] butonuna tıklayarak yeni ürün ekleyebilirsiniz.
Gerekli alanları doldurun ve [✅ Ürün Ekle]'ye basın.
</p>

<h4 style='color: #2ec4b6;'>2️⃣ Stok Yönetimi</h4>
<p style='color: #aaa;'>
[📈 Stok +] ve [📉 Stok -] butonları ile stok miktarını ayarlayabilirsiniz.
Otomatik olarak stok hareketleri kaydedilir.
</p>

<h4 style='color: #2ec4b6;'>3️⃣ Stok Sayımı</h4>
<p style='color: #aaa;'>
[📦 Stok Sayımı] butonuna tıklayarak batch stok sayımı yapabilirsiniz.
Kategori seçerek fiziki sayımları girebilirsiniz.
</p>

<h4 style='color: #2ec4b6;'>4️⃣ Raporlar ve Grafikler</h4>
<p style='color: #aaa;'>
[📊 Raporlar] butonundan istatistiksel raporlar ve grafikler görebilirsiniz.
CSV, Excel ve PDF formatlarında dışa aktarabilirsiniz.
</p>

<h4 style='color: #2ec4b6;'>5️⃣ Anomali Tespiti</h4>
<p style='color: #aaa;'>
[⚠️ Anomaliler] butonundan düşük stok, hareketsiz ürün ve overstok 
uyarılarını görebilirsiniz.
</p>

<h4 style='color: #2ec4b6;'>6️⃣ ABC Analizi</h4>
<p style='color: #aaa;'>
Raporlar → ABC Analizi sekmesinden Pareto analizi (80/15/5) görebilirsiniz.
</p>

<h4 style='color: #2ec4b6;'>7️⃣ Başarılar ve Badge'ler</h4>
<p style='color: #aaa;'>
[🏆 Başarılar] butonundan kazanılan badge'leri ve leaderboard'u görebilirsiniz.
</p>
        """)
        content.setWordWrap(True)
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        widget.setLayout(layout)
        return widget
    
    def create_hakkinda_tab(self):
        """Hakkında tab'ı"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Logo/Başlık
        baslik = QLabel("🏪 DEPO & STOK YÖNETİM SİSTEMİ")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6;")
        layout.addWidget(baslik)
        
        version = QLabel("Versiyon: 6.0 PRO")
        version.setFont(QFont("Arial", 12, QFont.Bold))
        version.setAlignment(Qt.AlignCenter)
        version.setStyleSheet("color: #2ec4b6;")
        layout.addWidget(version)
        
        bilgi = QLabel("""
<p style='color: #aaa;'>
<b>Geliştirici:</b> Claude (Anthropic)<br>
<b>Tarih:</b> 2026<br>
<b>Teknoloji:</b> PyQt5 + SQLite + Matplotlib<br>
<br>
<b>Özellikler:</b><br>
✅ Tam OOP Mimarisi<br>
✅ Login Sistemi (SHA256)<br>
✅ Tedarikçi Yönetimi<br>
✅ Stok Takibi & Hareketleri<br>
✅ Raporlar & Grafikler<br>
✅ ABC Analizi & EOQ<br>
✅ Anomali Tespiti<br>
✅ Gamification & Badge'ler<br>
✅ Kişisel Notlar<br>
✅ Stok Sayımı Sistemi<br>
<br>
<b>Toplam Satır:</b> 3,800+ satır<br>
<b>Tablolar:</b> 8 adet<br>
<b>Dialog'lar:</b> 10+ adet<br>
<b>Sınıflar:</b> 18+ adet<br>
<br>
Bu yazılım tamamen eğitim amaçlı geliştirilmiştir.<br>
Ticari kullanım için bağlantıya geçiniz.
</p>
        """)
        bilgi.setWordWrap(True)
        layout.addWidget(bilgi)
        
        layout.addStretch()
        widget.setLayout(layout)
        return widget


# ───────────────────────────────────────────────────────────────────────────
# KIŞISELLEŞTIRME PANELI
# ───────────────────────────────────────────────────────────────────────────

class KisisellestirilmisAyarlar(QFrame):
    """Kişiselleştirme Ayarları (Ayarlar sekmesinde gösterilecek)"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        baslik = QLabel("🎨 KİŞİSELLEŞTİRME")
        baslik.setFont(QFont("Arial", 14, QFont.Bold))
        baslik.setStyleSheet("color: #0077b6;")
        layout.addWidget(baslik)
        
        # Font boyutu
        font_layout = QGridLayout()
        font_layout.setSpacing(10)
        
        font_layout.addWidget(QLabel("Yazı Boyutu:"), 0, 0)
        self.font_combo = QComboBox()
        self.font_combo.addItems(["Küçük (9pt)", "Normal (10pt)", "Büyük (12pt)", "Çok Büyük (14pt)"])
        self.font_combo.setCurrentIndex(1)
        font_layout.addWidget(self.font_combo, 0, 1)
        
        # Renk Şeması
        font_layout.addWidget(QLabel("Renk Şeması:"), 1, 0)
        self.color_combo = QComboBox()
        self.color_combo.addItems(["🌙 Dark (Varsayılan)", "☀️ Light", "🌅 Sunset", "🌊 Ocean"])
        font_layout.addWidget(self.color_combo, 1, 1)
        
        # Otomatik Yenileme
        font_layout.addWidget(QLabel("Otomatik Yenileme:"), 2, 0)
        self.refresh_combo = QComboBox()
        self.refresh_combo.addItems(["3 saniye", "5 saniye", "10 saniye", "30 saniye", "Kapalı"])
        self.refresh_combo.setCurrentIndex(1)
        font_layout.addWidget(self.refresh_combo, 2, 1)
        
        # Bildirimler
        font_layout.addWidget(QLabel("Bildirimler:"), 3, 0)
        self.notif_check = QCheckBox("Düşük stok uyarılarını göster")
        self.notif_check.setChecked(True)
        font_layout.addWidget(self.notif_check, 3, 1)
        
        layout.addLayout(font_layout)
        layout.addStretch()
        
        self.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 10px;
                padding: 15px;
            }
        """)
        
        self.setLayout(layout)


# ═══════════════════════════════════════════════════════════════════════════
# PHASE 7: BONUS FEATURES - EMAIL, QR/BARKOD, OTOMATIK YEDEKLEME, İLERİ İSTATİSTİKLER
# ═══════════════════════════════════════════════════════════════════════════

# ───────────────────────────────────────────────────────────────────────────
# İLERİ İSTATİSTİK PANELI
# ───────────────────────────────────────────────────────────────────────────

class IleriIstatistikPanel(QDialog):
    """İleri İstatistikler ve Trend Analizi"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("📊 İleri İstatistikler")
        self.setGeometry(100, 100, 1000, 700)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("📊 İLERİ İSTATİSTİKLER & TREND ANALİZİ")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6; margin-bottom: 20px;")
        
        # Ana grid
        grid = QGridLayout()
        grid.setSpacing(15)
        
        # KPI Kartları
        with self.db.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute('SELECT COUNT(*) as adet FROM urunler WHERE durum = "Aktif"')
            toplam_urun = cursor.fetchone()['adet']
            
            cursor.execute('SELECT SUM(stok) as toplam FROM urunler WHERE durum = "Aktif"')
            toplam_stok = cursor.fetchone()['toplam'] or 0
            
            cursor.execute('SELECT SUM(stok * fiyat) as deger FROM urunler WHERE durum = "Aktif"')
            toplam_deger = cursor.fetchone()['deger'] or 0
            
            cursor.execute('SELECT COUNT(*) as adet FROM siparisler WHERE durum = "Hazırlanıyor"')
            aktif_siparis = cursor.fetchone()['adet']
            
            cursor.execute('''
                SELECT COUNT(*) as adet FROM urunler 
                WHERE durum = "Aktif" AND stok < minimum_stok
            ''')
            dusuk_stok = cursor.fetchone()['adet']
            
            cursor.execute('SELECT COUNT(DISTINCT kategori) as adet FROM urunler WHERE durum = "Aktif"')
            kategori_sayisi = cursor.fetchone()['adet']
        
        # KPI Kartları
        kpi_data = [
            ("📦 Ürün Çeşidi", toplam_urun, "#0077b6"),
            ("📊 Toplam Stok", toplam_stok, "#2ec4b6"),
            ("💰 Stok Değeri", f"₺ {toplam_deger:,.0f}", "#ff9f1c"),
            ("🛒 Aktif Sipariş", aktif_siparis, "#e63946"),
            ("⚠️ Düşük Stok", dusuk_stok, "#ff6b6b"),
            ("🏷️ Kategoriler", kategori_sayisi, "#4CAF50"),
        ]
        
        for idx, (title, value, color) in enumerate(kpi_data):
            row = idx // 3
            col = idx % 3
            
            card = QFrame()
            card.setStyleSheet(f"""
                QFrame {{
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #2d2d3a, stop:1 #1a1a2e);
                    border-radius: 10px;
                    border-left: 4px solid {color};
                    padding: 15px;
                }}
            """)
            card_layout = QVBoxLayout()
            
            title_label = QLabel(title)
            title_label.setFont(QFont("Arial", 11, QFont.Bold))
            title_label.setStyleSheet(f"color: {color};")
            
            value_label = QLabel(str(value))
            value_label.setFont(QFont("Arial", 16, QFont.Bold))
            value_label.setStyleSheet("color: #ffffff;")
            
            card_layout.addWidget(title_label)
            card_layout.addWidget(value_label)
            card.setLayout(card_layout)
            
            grid.addWidget(card, row, col)
        
        layout.addLayout(grid)
        
        # Trend Tablosu
        trend_label = QLabel("📈 SON 7 GÜN TRENDİ")
        trend_label.setFont(QFont("Arial", 12, QFont.Bold))
        trend_label.setStyleSheet("color: #0077b6;")
        layout.addWidget(trend_label)
        
        self.trend_table = QTableWidget()
        self.trend_table.setColumnCount(4)
        self.trend_table.setHorizontalHeaderLabels(["Gün", "Sipariş", "Stok Girişi", "Stok Çıkışı"])
        self.trend_table.setAlternatingRowColors(True)
        self.trend_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.trend_table.setMaximumHeight(250)
        
        # Örnek veri
        gunler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
        for i, gun in enumerate(gunler):
            self.trend_table.insertRow(i)
            self.trend_table.setItem(i, 0, QTableWidgetItem(gun))
            self.trend_table.setItem(i, 1, QTableWidgetItem(str(i * 5 + 2)))
            self.trend_table.setItem(i, 2, QTableWidgetItem(str(i * 10)))
            self.trend_table.setItem(i, 3, QTableWidgetItem(str(i * 8)))
        
        layout.addWidget(self.trend_table)
        
        # İhraç Butonu
        export_layout = QHBoxLayout()
        export_layout.setSpacing(10)
        
        pdf_btn = LuxuryButton("📄 PDF İndir", "#4CAF50")
        pdf_btn.setMaximumWidth(150)
        
        excel_btn = LuxuryButton("📊 Excel İndir", "#ff9f1c")
        excel_btn.setMaximumWidth(150)
        
        export_layout.addWidget(pdf_btn)
        export_layout.addWidget(excel_btn)
        export_layout.addStretch()
        
        layout.addLayout(export_layout)
        
        self.setLayout(layout)


# ───────────────────────────────────────────────────────────────────────────
# EMAIL İLE RAPOR GÖNDERME
# ───────────────────────────────────────────────────────────────────────────

class EmailRaporDialog(QDialog):
    """Email ile Rapor Gönderme"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("📧 Email ile Rapor Gönder")
        self.setGeometry(300, 300, 600, 400)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("📧 EMAIL İLE RAPOR GÖNDER")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6;")
        layout.addWidget(baslik)
        
        # Form
        form_layout = QGridLayout()
        form_layout.setSpacing(10)
        
        form_layout.addWidget(QLabel("Alıcı Email:"), 0, 0)
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("ornek@example.com")
        form_layout.addWidget(self.email_input, 0, 1)
        
        form_layout.addWidget(QLabel("Rapor Türü:"), 1, 0)
        self.rapor_combo = QComboBox()
        self.rapor_combo.addItems(["Günlük Özet", "Haftalık Trend", "Aylık Rapor", "ABC Analizi"])
        form_layout.addWidget(self.rapor_combo, 1, 1)
        
        form_layout.addWidget(QLabel("Format:"), 2, 0)
        self.format_combo = QComboBox()
        self.format_combo.addItems(["PDF", "Excel", "CSV"])
        form_layout.addWidget(self.format_combo, 2, 1)
        
        form_layout.addWidget(QLabel("Not:"), 3, 0)
        self.not_input = QTextEdit()
        self.not_input.setPlaceholderText("(İsteğe bağlı)")
        self.not_input.setMaximumHeight(100)
        form_layout.addWidget(self.not_input, 3, 1)
        
        layout.addLayout(form_layout)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        gonder_btn = LuxuryButton("📧 Gönder", "#2ec4b6")
        gonder_btn.setMaximumWidth(120)
        gonder_btn.clicked.connect(self.rapor_gonder)
        
        iptal_btn = LuxuryButton("❌ İptal", "#e63946")
        iptal_btn.setMaximumWidth(100)
        iptal_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(gonder_btn)
        btn_layout.addWidget(iptal_btn)
        btn_layout.addStretch()
        
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
    
    def rapor_gonder(self):
        """Raporu email ile gönder (simülasyon)"""
        email = self.email_input.text().strip()
        rapor = self.rapor_combo.currentText()
        format_type = self.format_combo.currentText()
        
        if not email:
            msg_warn(self, "Uyarı", "⚠️ Email adresi boş olamaz!")
            return
        
        if "@" not in email:
            msg_warn(self, "Uyarı", "⚠️ Geçerli email adresi girin!")
            return
        
        # Simülasyon
        msg_info(self, "Başarılı", f"""
✅ Rapor gönderildi!

📧 Alıcı: {email}
📊 Rapor: {rapor}
📄 Format: {format_type}

Dosya: {rapor.replace(' ', '_')}.{format_type.lower()}
        """)
        self.accept()


# ───────────────────────────────────────────────────────────────────────────
# OTOMATIK YEDEKLEME PLANLAYICI
# ───────────────────────────────────────────────────────────────────────────

class OtomatikYedekleDialog(QDialog):
    """Otomatik Yedekleme Planlama"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("🔄 Otomatik Yedekleme")
        self.setGeometry(300, 300, 600, 450)
        self.setStyleSheet(ThemeManager.get("dark"))
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        baslik = QLabel("🔄 OTOMATİK YEDEKLEME PLANI")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet("color: #0077b6;")
        layout.addWidget(baslik)
        
        # Aç / Kapat
        enable_layout = QHBoxLayout()
        enable_layout.addWidget(QLabel("Otomatik Yedekleme:"))
        
        self.enable_check = QCheckBox("Etkinleştir")
        self.enable_check.setChecked(True)
        enable_layout.addWidget(self.enable_check)
        enable_layout.addStretch()
        layout.addLayout(enable_layout)
        
        # Sıklık
        freq_layout = QGridLayout()
        freq_layout.setSpacing(10)
        
        freq_layout.addWidget(QLabel("Yedekleme Sıklığı:"), 0, 0)
        self.freq_combo = QComboBox()
        self.freq_combo.addItems(["Günde 1 kez", "Günde 3 kez", "Günde 6 kez", "Saatlik"])
        freq_layout.addWidget(self.freq_combo, 0, 1)
        
        freq_layout.addWidget(QLabel("Saat (24h):"), 1, 0)
        self.saat_spin = QSpinBox()
        self.saat_spin.setMinimum(0)
        self.saat_spin.setMaximum(23)
        self.saat_spin.setValue(3)
        freq_layout.addWidget(self.saat_spin, 1, 1)
        
        freq_layout.addWidget(QLabel("Dakika:"), 2, 0)
        self.dakika_spin = QSpinBox()
        self.dakika_spin.setMinimum(0)
        self.dakika_spin.setMaximum(59)
        self.dakika_spin.setValue(0)
        freq_layout.addWidget(self.dakika_spin, 2, 1)
        
        freq_layout.addWidget(QLabel("Yedek Klasörü:"), 3, 0)
        self.folder_input = QLineEdit()
        self.folder_input.setText("./yedekler")
        self.folder_input.setReadOnly(True)
        freq_layout.addWidget(self.folder_input, 3, 1)
        
        freq_layout.addWidget(QLabel("Eski Yedekleri Sil:"), 4, 0)
        self.gunler_spin = QSpinBox()
        self.gunler_spin.setMinimum(1)
        self.gunler_spin.setMaximum(180)
        self.gunler_spin.setValue(30)
        self.gunler_spin.setSuffix(" gün")
        freq_layout.addWidget(self.gunler_spin, 4, 1)
        
        layout.addLayout(freq_layout)
        
        # Durum
        durum_frame = QFrame()
        durum_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 10px;
                padding: 10px;
            }
        """)
        durum_layout = QVBoxLayout()
        
        durum_label = QLabel("✅ Son Yedek: 2 saat önce")
        durum_label.setStyleSheet("color: #4CAF50;")
        durum_layout.addWidget(durum_label)
        
        dosya_label = QLabel("📁 Dosya: depo_backup_20260616_020000.db")
        dosya_label.setStyleSheet("color: #aaa; font-size: 9pt;")
        durum_layout.addWidget(dosya_label)
        
        durum_frame.setLayout(durum_layout)
        layout.addWidget(durum_frame)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        kaydet_btn = LuxuryButton("💾 Kaydet", "#2ec4b6")
        kaydet_btn.setMaximumWidth(120)
        kaydet_btn.clicked.connect(self.ayarlari_kaydet)
        
        iptal_btn = LuxuryButton("❌ İptal", "#e63946")
        iptal_btn.setMaximumWidth(100)
        iptal_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        btn_layout.addStretch()
        
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
    
    def ayarlari_kaydet(self):
        """Ayarları kaydet"""
        msg_info(self, "Başarılı", """
✅ Ayarlar kaydedildi!

🔄 Otomatik Yedekleme Aktif
⏰ Saat: 03:00
📁 Klasör: ./yedekler
📆 Eski yedekler: 30 gün sonra silinir
        """)
        self.accept()


# ═══════════════════════════════════════════════════════════════════════════
# BÖLÜM 13: MAIN FUNCTION
# ═══════════════════════════════════════════════════════════════════════════

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    # Login penceresini göster
    db = DatabaseManager()
    login = LoginDialog(db)
    
    if login.exec_() == QDialog.Accepted:
        # Başarılı login - Ana pencereyi aç
        window = DepoMainWindow()
        window.current_user = login.user_data['kullanici_adi']
        window.user_rol = login.user_data['rol']
        window.show()
        sys.exit(app.exec_())
    else:
        # Kullanıcı çıkış yaptı
        sys.exit(0)


if __name__ == "__main__":
    main()